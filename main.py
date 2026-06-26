"""
API del optimizador WFM — backend para Render.
Expone un endpoint POST /generar-roster que:
  1. Lee el histórico (CSV en Supabase Storage o enviado en la petición)
  2. Corre el optimizador (motor.py) + repartidor (repartidor.py)
  3. Genera breaks por país
  4. Escribe asignaciones + breaks en Supabase

Seguridad: protegido por una API key secreta (header X-API-Key).
"""
import os
import io
from datetime import datetime, timedelta
from collections import defaultdict

import pandas as pd
from fastapi import FastAPI, Header, HTTPException, UploadFile, File, Form
from fastapi.middleware.cors import CORSMiddleware
from sqlalchemy import create_engine, text

import motor
import repartidor

app = FastAPI(title="WFM Optimizer API")

# CORS: permitir que skyewfm.com llame a esta API
app.add_middleware(
    CORSMiddleware,
    allow_origins=["https://skyewfm.com", "https://www.skyewfm.com"],
    allow_methods=["*"],
    allow_headers=["*"],
)

# --- Credenciales desde variables de entorno (se configuran en Render) ---
DB_URL = os.environ.get("DB_URL")            # cadena de conexión del pooler de Supabase
API_KEY = os.environ.get("API_KEY")          # clave secreta para proteger el endpoint

LIBRES = {p: {p, (p + 1) % 7} for p in range(7)}


def get_engine():
    if not DB_URL:
        raise HTTPException(500, "DB_URL no configurada en el servidor")
    return create_engine(DB_URL, pool_pre_ping=True)


def check_key(x_api_key):
    if not API_KEY or x_api_key != API_KEY:
        raise HTTPException(401, "API key inválida")


# ---------- breaks por país ----------
def _t(hhmm):
    h, m = map(int, str(hhmm).split(":")[:2])
    return datetime(2000, 1, 1, h % 24, m)


def breaks_colombia(hi, dur, idx, ventana=30):
    ini = _t(hi); off = timedelta(minutes=(idx * 5) % ventana); r = []
    r.append(((ini + timedelta(hours=2) + off).strftime("%H:%M"), 10, "break"))
    r.append(((ini + timedelta(hours=dur / 2) + off).strftime("%H:%M"), 30, "almuerzo"))
    r.append(((ini + timedelta(hours=dur - 2) + off).strftime("%H:%M"), 10, "break"))
    return r


def breaks_espana(hi, dur, idx, ventana=30):
    ini = _t(hi); off = timedelta(minutes=(idx * 5) % ventana); r = []
    r.append(((ini + timedelta(hours=dur / 2) + off).strftime("%H:%M"), 30, "descanso"))
    for h in range(int(dur)):
        r.append(((ini + timedelta(hours=h, minutes=50)).strftime("%H:%M"), 5, "break"))
    return r


def dur_turno(hi, hf):
    a = _t(hi); b = _t(hf); d = (b - a).seconds / 3600
    return d if d > 0 else d + 24


@app.get("/")
def health():
    return {"status": "ok", "service": "WFM Optimizer API"}


@app.get("/capacidad-anual")
def capacidad_anual(
    x_api_key: str = Header(None),
    anio: int = 2026,
    scope: str = "Nacional España",
    K: int = 4,
    aht: int = 420, sla: float = 0.80, asa: int = 20, occ: float = 0.65,
    utl: float = 0.88, esp_max: int = 12, largo: int = 9,
    nda_obj: float = 0.96, paciencia: int = 90,
    estructura: str = "mixto", absentismo: float = 0.15,
    plantilla_actual: int = 129,
    campana: str = None,
):
    """Dimensiona los 12 meses del año: agentes necesarios por mes."""
    check_key(x_api_key)
    import math as _m
    engine = get_engine()
    cid = _id_campana(engine, campana)
    file_bytes = _historico_a_csv_bytes(engine, cid)
    # fechas con datos reales (para marcar estado)
    if cid is not None:
        dch = pd.read_sql(text("SELECT DISTINCT fecha FROM historico_llamadas WHERE campana_id=:c"), engine, params={"c": cid})
    else:
        dch = pd.read_sql(text("SELECT DISTINCT fecha FROM historico_llamadas"), engine)
    dias_datos = set(pd.to_datetime(dch["fecha"]).dt.normalize())

    MESES = ["Ene","Feb","Mar","Abr","May","Jun","Jul","Ago","Sep","Oct","Nov","Dic"]
    filas = []
    for mth in range(1, 13):
        mes_str = f"{anio}-{mth:02d}"
        ini = pd.Timestamp(anio, mth, 1); dm = pd.date_range(ini, ini + pd.offsets.MonthEnd(0))
        nr = sum(1 for t in dm if t in dias_datos)
        estado = "Real" if nr >= len(dm) else ("En curso" if nr > 0 else "Proyectado")
        try:
            largo_df = motor.largo_desde_historico(file_bytes, mes_str, scope, K, 6, 0.0, mixto=False)
            Sx = motor.dimension_roster(largo_df, aht, sla, asa, occ, utl, esp_max, largo,
                                        nda_obj, paciencia, estructura)
            presentes = Sx["te"] + Sx["tc"]
            nomina = _m.ceil(presentes / (1 - absentismo)) if absentismo < 1 else presentes
            filas.append({"mes": MESES[mth-1], "estado": estado, "volumen": Sx["total"],
                          "espana": Sx["te"], "colombia": Sx["tc"], "presentes": presentes,
                          "en_nomina": nomina})
        except Exception:
            filas.append({"mes": MESES[mth-1], "estado": estado, "volumen": 0,
                          "espana": 0, "colombia": 0, "presentes": 0, "en_nomina": 0})

    validos = [f for f in filas if f["en_nomina"] > 0]
    pico = max(validos, key=lambda f: f["en_nomina"]) if validos else None
    prom = round(sum(f["en_nomina"] for f in validos) / len(validos)) if validos else 0
    return {
        "ok": True, "anio": anio,
        "tabla": filas,
        "mes_pico": {"mes": pico["mes"], "en_nomina": pico["en_nomina"]} if pico else None,
        "promedio_nomina": prom,
        "plantilla_actual": plantilla_actual,
    }


@app.get("/proyeccion-anual")
def proyeccion_anual(
    x_api_key: str = Header(None),
    anio: int = None,
    K: int = 4,
    recencia: float = 0.0,
    scope: str = "Nacional España",
    colas: str = None,
    ajuste_meses: str = None,   # 12 valores % separados por coma, ej "0,0,10,0,..."
    campana: str = None,
):
    """Proyección anual de volumen: Base (sin ajustes) vs Escenario (con ajustes)."""
    check_key(x_api_key)
    import holidays as _hol
    engine = get_engine()
    cid = _id_campana(engine, campana)
    if cid is not None:
        d = pd.read_sql(text("SELECT fecha, cola, entrantes FROM historico_llamadas WHERE campana_id=:c"), engine, params={"c": cid})
    else:
        d = pd.read_sql(text("SELECT fecha, cola, entrantes FROM historico_llamadas"), engine)
    if d.empty:
        return {"ok": True, "vacio": True}
    d["fecha"] = pd.to_datetime(d["fecha"]).dt.normalize()
    d["entrantes"] = pd.to_numeric(d["entrantes"], errors="coerce").fillna(0)

    share = d.groupby("cola")["entrantes"].sum().sort_values(ascending=False)
    share_pct = (share / share.sum() * 100).round(1)

    colas_sel = [c.strip() for c in colas.split(",")] if colas else list(share.index)
    colas_sel = [c for c in colas_sel if c in share.index] or list(share.index)
    d = d[d["cola"].isin(colas_sel)]

    dia = d.groupby("fecha")["entrantes"].sum()
    if dia.empty:
        return {"ok": True, "vacio": False, "sin_datos": True}
    anios_disp = sorted({t.year for t in dia.index})
    if anio is None:
        anio = anios_disp[-1]

    adj_mes = [0.0] * 12
    if ajuste_meses:
        parts = [p.strip() for p in ajuste_meses.split(",")]
        for i, p in enumerate(parts[:12]):
            try: adj_mes[i] = float(p) / 100.0
            except: adj_mes[i] = 0.0

    # peso por cola (sin ajuste de cola individual por ahora; el ajuste va por mes)
    cola_factor = 1.0

    if "Catal" in scope:
        ES = _hol.Spain(years=range(min(anios_disp) - 1, max(anios_disp) + 2), subdiv="CT")
    else:
        ES = _hol.Spain(years=range(min(anios_disp) - 1, max(anios_disp) + 2))

    def fest(t): return t.date() in ES

    idx_dow = dia.index.dayofweek
    def proj_dia(t):
        wd = 6 if fest(t) else t.dayofweek
        s = dia[(dia.index < t) & (idx_dow == wd)]
        if wd != 6:
            s = s[[not fest(x) for x in s.index]]
        s = s.tail(K)
        if len(s) == 0: return 0.0
        v = s.values[::-1]
        import numpy as _np
        w = _np.array([(1 - recencia) ** j for j in range(len(v))])
        return float((v * w).sum() / w.sum())

    MESES = ["Ene","Feb","Mar","Abr","May","Jun","Jul","Ago","Sep","Oct","Nov","Dic"]
    filas = []
    for mth in range(1, 13):
        ini = pd.Timestamp(year=anio, month=mth, day=1); fin = ini + pd.offsets.MonthEnd(0)
        dias_mes = pd.date_range(ini, fin)
        real_part = pbase = 0.0
        for t in dias_mes:
            if t in dia.index: real_part += float(dia.loc[t])
            else: pbase += proj_dia(t)
        n_real = sum(1 for t in dias_mes if t in dia.index)
        estado = "REAL" if n_real >= len(dias_mes) else ("EN CURSO" if n_real > 0 else "PROYECTADO")
        f_mes = cola_factor * (1 + adj_mes[mth - 1])
        base_tot = real_part + pbase
        scn_tot = real_part + pbase * f_mes
        delta = (scn_tot / base_tot - 1) * 100 if base_tot > 0 else 0.0
        filas.append({"mes": MESES[mth-1], "estado": estado, "real": int(round(real_part)),
                      "base": int(round(base_tot)), "escenario": int(round(scn_tot)), "delta": round(delta,1)})

    # Promedios históricos
    pdow = dia.groupby(idx_dow).mean().reindex(range(7)).fillna(0)
    NOM7 = ["Lun","Mar","Mié","Jue","Vie","Sáb","Dom"]
    prom_dow = [{"dia": NOM7[i], "promedio": round(float(pdow.values[i]),1)} for i in range(7)]
    dia_anio = dia[dia.index.year == anio]
    sem = dia_anio.resample("W").sum()
    por_semana = [{"semana": idx.strftime("%Y-%m-%d"), "entrantes": int(v)} for idx, v in sem.items()]

    tb = sum(f["base"] for f in filas); tsc = sum(f["escenario"] for f in filas)
    return {
        "ok": True, "vacio": False,
        "anio": anio, "anios_disponibles": anios_disp,
        "colas_disponibles": [{"cola": c, "pct": float(share_pct[c])} for c in share.index],
        "tabla": filas,
        "totales": {"base": tb, "escenario": tsc, "diferencia": tsc - tb,
                    "delta_pct": round((tsc/tb - 1)*100, 1) if tb else None},
        "promedio_dow": prom_dow,
        "por_semana": por_semana,
    }


@app.get("/dashboard-historico")
def dashboard_historico(
    x_api_key: str = Header(None),
    anio: int = None,
    mes: int = None,
    semana: int = None,
    colas: str = None,   # colas separadas por coma; vacío = todas
    dows: str = None,    # días de semana 0-6 separados por coma; vacío = todos
    campana: str = None, # nombre de campaña; vacío = todas
):
    """Devuelve datos agregados del histórico para el dashboard, según filtros."""
    check_key(x_api_key)
    engine = get_engine()
    cid = _id_campana(engine, campana)
    if cid is not None:
        df = pd.read_sql(text("SELECT fecha, hora, cola, entrantes, atendidas, abandonadas "
                              "FROM historico_llamadas WHERE campana_id=:c"),
                         engine, params={"c": cid})
    else:
        df = pd.read_sql(text("SELECT fecha, hora, cola, entrantes, atendidas, abandonadas FROM historico_llamadas"), engine)
    if df.empty:
        return {"ok": True, "vacio": True}
    df["fecha"] = pd.to_datetime(df["fecha"])
    df["anio"] = df["fecha"].dt.year
    df["mes"] = df["fecha"].dt.month
    df["semana"] = df["fecha"].dt.isocalendar().week.astype(int)
    df["dow"] = df["fecha"].dt.dayofweek

    if anio:
        df = df[df["anio"] == anio]
    if mes:
        df = df[df["mes"] == mes]
    if semana:
        df = df[df["semana"] == semana]
    if colas:
        lista = [c.strip() for c in colas.split(",") if c.strip()]
        if lista:
            df = df[df["cola"].isin(lista)]
    if dows:
        lista_d = [int(x) for x in dows.split(",") if x.strip().isdigit()]
        if lista_d:
            df = df[df["dow"].isin(lista_d)]

    if df.empty:
        return {"ok": True, "vacio": False, "sin_datos_filtro": True}

    ent = int(df["entrantes"].sum()); at = int(df["atendidas"].sum()); ab = int(df["abandonadas"].sum())

    # Por cola
    by = df.groupby("cola")[["entrantes", "atendidas", "abandonadas"]].sum().reset_index()
    by["pat"] = (by["atendidas"] / by["entrantes"].replace(0, pd.NA) * 100).round(1)
    by = by.sort_values("entrantes", ascending=False)
    por_cola = by.to_dict("records")

    # % atención en el tiempo (por fecha)
    serie = df.groupby(df["fecha"].dt.strftime("%Y-%m-%d"))[["entrantes", "atendidas"]].sum()
    serie["pat"] = (serie["atendidas"] / serie["entrantes"].replace(0, pd.NA) * 100).round(1)
    en_tiempo = [{"fecha": idx, "pat": (None if pd.isna(row["pat"]) else float(row["pat"]))}
                 for idx, row in serie.iterrows()]

    # Por franja horaria
    fr = df.groupby("hora")[["entrantes", "atendidas", "abandonadas"]].sum()
    fr["pat"] = (fr["atendidas"] / fr["entrantes"].replace(0, pd.NA) * 100).round(1)
    por_hora = [{"hora": int(h), "entrantes": int(row["entrantes"]),
                 "pat": (None if pd.isna(row["pat"]) else float(row["pat"]))}
                for h, row in fr.iterrows()]

    return {
        "ok": True, "vacio": False,
        "metricas": {"entrantes": ent, "atendidas": at, "abandonadas": ab,
                     "pct_atencion": round(at / ent * 100, 1) if ent else None},
        "por_cola": por_cola,
        "en_tiempo": en_tiempo,
        "por_hora": por_hora,
    }


@app.get("/dashboard-opciones")
def dashboard_opciones(x_api_key: str = Header(None), campana: str = None):
    """Devuelve los valores disponibles para los filtros (años, colas)."""
    check_key(x_api_key)
    engine = get_engine()
    cid = _id_campana(engine, campana)
    if cid is not None:
        df = pd.read_sql(text("SELECT DISTINCT fecha, cola FROM historico_llamadas WHERE campana_id=:c"),
                         engine, params={"c": cid})
    else:
        df = pd.read_sql(text("SELECT DISTINCT fecha, cola FROM historico_llamadas"), engine)
    if df.empty:
        return {"ok": True, "anios": [], "colas": []}
    df["fecha"] = pd.to_datetime(df["fecha"])
    anios = sorted(df["fecha"].dt.year.unique().tolist())
    colas = sorted(df["cola"].unique().tolist())
    return {"ok": True, "anios": anios, "colas": colas}


@app.post("/historico/actualizar")
async def historico_actualizar(
    x_api_key: str = Header(None),
    campana: str = Form("Endesa"),
    archivo: UploadFile = File(...),
):
    """
    Sube un CSV y lo fusiona con el histórico acumulado de UNA campaña (upsert):
    - Filas con fecha+hora+cola ya existentes en esa campaña: se SOBRESCRIBEN.
    - Filas nuevas: se AGREGAN.
    Columnas esperadas: fecha, hora, cola, entrantes, atendidas, abandonadas.
    """
    check_key(x_api_key)
    engine = get_engine()
    cid = _id_campana(engine, campana)
    if cid is None:
        raise HTTPException(400, f"La campaña '{campana}' no existe.")
    file_bytes = await archivo.read()
    df = pd.read_csv(io.BytesIO(file_bytes))
    df.columns = [c.strip().lower() for c in df.columns]
    req = {"fecha", "hora", "cola", "entrantes", "atendidas", "abandonadas"}
    falta = req - set(df.columns)
    if falta:
        raise HTTPException(400, f"Faltan columnas en el CSV: {falta}")
    df = df[list(req)].copy()
    df["fecha"] = pd.to_datetime(df["fecha"]).dt.date
    for c in ["hora", "entrantes", "atendidas", "abandonadas"]:
        df[c] = df[c].astype(int)

    filas_csv = len(df)
    # upsert por lotes usando ON CONFLICT
    insertadas = 0
    with engine.begin() as conn:
        for _, r in df.iterrows():
            conn.execute(text("""
                INSERT INTO historico_llamadas (fecha, hora, cola, entrantes, atendidas, abandonadas, campana_id, actualizado_en)
                VALUES (:f, :h, :c, :e, :a, :ab, :cid, now())
                ON CONFLICT (fecha, hora, cola, campana_id)
                DO UPDATE SET entrantes=EXCLUDED.entrantes, atendidas=EXCLUDED.atendidas,
                              abandonadas=EXCLUDED.abandonadas, actualizado_en=now()
            """), {"f": r["fecha"], "h": int(r["hora"]), "c": str(r["cola"]),
                   "e": int(r["entrantes"]), "a": int(r["atendidas"]), "ab": int(r["abandonadas"]),
                   "cid": cid})
            insertadas += 1

    # rango del histórico de esa campaña tras la actualización
    rango = pd.read_sql(text("SELECT MIN(fecha) AS desde, MAX(fecha) AS hasta, COUNT(*) AS filas "
                             "FROM historico_llamadas WHERE campana_id=:c"), engine, params={"c": cid})
    return {
        "ok": True,
        "filas_csv": filas_csv,
        "filas_procesadas": insertadas,
        "historico_desde": str(rango["desde"][0]),
        "historico_hasta": str(rango["hasta"][0]),
        "total_filas_historico": int(rango["filas"][0]),
    }


@app.get("/historico/estado")
def historico_estado(x_api_key: str = Header(None), campana: str = None):
    """Devuelve hasta qué fecha hay histórico cargado (de la campaña indicada o de todas)."""
    check_key(x_api_key)
    engine = get_engine()
    cid = _id_campana(engine, campana)
    if cid is not None:
        rango = pd.read_sql(text("SELECT MIN(fecha) AS desde, MAX(fecha) AS hasta, COUNT(*) AS filas "
                                 "FROM historico_llamadas WHERE campana_id=:c"), engine, params={"c": cid})
    else:
        rango = pd.read_sql(text("SELECT MIN(fecha) AS desde, MAX(fecha) AS hasta, COUNT(*) AS filas FROM historico_llamadas"), engine)
    if rango["filas"][0] == 0:
        return {"ok": True, "vacio": True}
    return {
        "ok": True, "vacio": False,
        "historico_desde": str(rango["desde"][0]),
        "historico_hasta": str(rango["hasta"][0]),
        "total_filas": int(rango["filas"][0]),
    }


def _id_campana(engine, campana):
    """Devuelve el id de una campaña por nombre (o None si no se pasa nombre)."""
    if not campana:
        return None
    row = pd.read_sql(text("SELECT id FROM campanas WHERE nombre=:n"), engine, params={"n": campana})
    return int(row["id"][0]) if not row.empty else None


def _historico_a_csv_bytes(engine, campana_id=None):
    """Lee el histórico de la base (filtrado por campaña si se indica) y lo devuelve como CSV en bytes."""
    if campana_id is not None:
        df = pd.read_sql(text("SELECT fecha, hora, cola, entrantes, atendidas, abandonadas "
                              "FROM historico_llamadas WHERE campana_id=:c ORDER BY fecha, hora"),
                         engine, params={"c": int(campana_id)})
    else:
        df = pd.read_sql(text("SELECT fecha, hora, cola, entrantes, atendidas, abandonadas "
                              "FROM historico_llamadas ORDER BY fecha, hora"), engine)
    return df.to_csv(index=False).encode()


# ============================================================
#  PLANEACIÓN · Carga GENÉRICA de histórico, guiada por config
#  por campaña (tabla campana_cargas). Soporta:
#    - formato 'intervalo'    (datos por intervalo 60/30/15 -> agrega a hora)
#    - formato 'diario_curva' (volumen diario + curva -> expande a 24 h)
#  Colas configurables (1..N) con reglas 'incluye'. Escribe en
#  historico_llamadas a granularidad HORARIA. Sin código por campaña.
#  (Helpers de curva/columnas/reparto compartidos por ambos formatos.)
# ============================================================
import unicodedata as _ud

_DIAS_CURVA = ["Lunes", "Martes", "Miércoles", "Jueves", "Viernes", "Sábado", "Domingo"]


def _norm_txt(s):
    """minúsculas, sin tildes, sin espacios extra (para comparar nombres de columna)."""
    s = str(s).strip().lower()
    return "".join(c for c in _ud.normalize("NFKD", s) if not _ud.combining(c))


def _col_por_nombre(columnas, *candidatos):
    """Nombre real de la 1ª columna que coincide (exacta o por 'contiene', normalizada)."""
    norm = {c: _norm_txt(c) for c in columnas}
    for cand in candidatos:
        cn = _norm_txt(cand)
        for real, n in norm.items():
            if n == cn:
                return real
        for real, n in norm.items():
            if cn in n:
                return real
    return None


def _curva_intradia_desde_excel(file_bytes, hoja="Curvas"):
    """
    Lee la curva intradía: localiza la fila de cabecera que contiene 'Intervalo',
    luego una columna por día (Lunes..Domingo) y filas de intervalo 0..23.
    Devuelve dict weekday(0=Lun..6=Dom) -> lista de 24 floats.
    """
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    if hoja not in wb.sheetnames:
        raise HTTPException(400, f"No encuentro la hoja de curva '{hoja}'. Hojas: {wb.sheetnames}")
    ws = wb[hoja]
    filas = list(ws.iter_rows(values_only=True))
    hi = None
    for i, r in enumerate(filas):
        if "intervalo" in [_norm_txt(x) for x in r if x is not None]:
            hi = i
            break
    if hi is None:
        raise HTTPException(400, f"No encuentro la cabecera 'Intervalo' en la hoja '{hoja}'.")
    header = filas[hi]
    idx_int = None
    idx_dia = {}
    for ci, c in enumerate(header):
        n = _norm_txt(c)
        if n == "intervalo":
            idx_int = ci
        for d, nom in enumerate(_DIAS_CURVA):
            if n == _norm_txt(nom):
                idx_dia[d] = ci
    if idx_int is None or len(idx_dia) < 7:
        raise HTTPException(400, f"La curva '{hoja}' no tiene 'Intervalo' + 7 días. Cabecera vista: {list(header)}")
    curva = {d: [0.0] * 24 for d in range(7)}
    for r in filas[hi + 1:]:
        try:
            h = int(r[idx_int])
        except (TypeError, ValueError):
            continue
        if 0 <= h <= 23:
            for d in range(7):
                v = r[idx_dia[d]]
                curva[d][h] = float(v) if v is not None else 0.0
    return curva


def _reparte_entero(total, pesos):
    """Reparte 'total' (entero) según 'pesos' preservando la suma exacta (mayor resto)."""
    if total <= 0:
        return [0] * len(pesos)
    crudos = [total * w for w in pesos]
    base = [int(x) for x in crudos]
    resto = int(round(total - sum(base)))
    orden = sorted(range(len(pesos)), key=lambda i: (crudos[i] - base[i]), reverse=True)
    for k in range(resto):
        base[orden[k % len(orden)]] += 1
    return base


def _leer_config_carga(engine, cid, tipo_carga="planeacion"):
    """Lee la config de carga (jsonb) de una campaña, o None si no existe."""
    tipo_carga = (tipo_carga or "planeacion").strip().lower()
    row = pd.read_sql(text("SELECT config FROM campana_cargas WHERE campana_id=:c AND lower(tipo_carga)=:t"),
                      engine, params={"c": cid, "t": tipo_carga})
    if row.empty:
        return None
    import json
    cfg = row["config"][0]
    if isinstance(cfg, str):
        cfg = json.loads(cfg)
    return cfg


def _leer_tabla(file_bytes, filename, hoja=None):
    """Lee el archivo a DataFrame: CSV si termina en .csv, si no Excel (hoja indicada o la 1ª)."""
    nombre = (filename or "").lower()
    if nombre.endswith(".csv"):
        return pd.read_csv(io.BytesIO(file_bytes))
    return pd.read_excel(io.BytesIO(file_bytes), sheet_name=(hoja if hoja else 0), engine="openpyxl")


def _a_hora_de_intervalo(valor, granularidad_min):
    """
    Convierte el valor de la columna de intervalo a una HORA 0-23 (para agregar a hora).
    Acepta: 'HH:MM[:SS]' (o un time de Excel, que se vuelve texto con ':'),
            un entero 0-23 cuando la granularidad es 60,
            o un ÍNDICE de intervalo (0..N-1) -> (idx*gran)//60.
    """
    if valor is None:
        return None
    s = str(valor).strip()
    if not s or s.lower() == "nan":
        return None
    if ":" in s:
        try:
            p = s.split(":")
            return (int(p[0]) * 60 + int(float(p[1]))) // 60 % 24
        except Exception:
            return None
    try:
        x = float(s)
    except Exception:
        return None
    g = int(granularidad_min or 60)
    if g >= 60 and 0 <= x <= 23:
        return int(x)
    return (int(round(x)) * g // 60) % 24


def _filtra_por_grupo(df, col_grupo_real, incluye):
    """Filtra el DataFrame a los valores 'incluye' de la columna de grupo. incluye=None -> todo."""
    if not col_grupo_real or incluye is None:
        return df
    vals = {_norm_txt(x) for x in incluye}
    return df[df[col_grupo_real].map(_norm_txt).isin(vals)]


def _in_clause(prefijo, valores):
    """Construye '(:p0, :p1, ...)' + sus params (IN portable Postgres/SQLite)."""
    ph = ", ".join(f":{prefijo}{i}" for i in range(len(valores)))
    pr = {f"{prefijo}{i}": v for i, v in enumerate(valores)}
    return "(" + ph + ")", pr


@app.get("/campanas")
def campanas_list(x_api_key: str = Header(None)):
    """Lista las campañas (para el selector del asistente)."""
    check_key(x_api_key)
    engine = get_engine()
    df = pd.read_sql(text("SELECT id, nombre, pais_tipo FROM campanas ORDER BY nombre"), engine)
    return {"ok": True, "campanas": df.to_dict(orient="records")}


@app.post("/campanas")
async def campanas_crear(
    x_api_key: str = Header(None),
    nombre: str = Form(...),
    pais_tipo: str = Form("MIXTA"),
):
    """Crea una campaña (o la devuelve si ya existe). Usado por el Paso 0 del asistente."""
    check_key(x_api_key)
    engine = get_engine()
    nom = (nombre or "").strip()
    if not nom:
        raise HTTPException(400, "El nombre de la campaña no puede estar vacío.")
    pt = (pais_tipo or "MIXTA").strip().upper()
    if pt not in ("ES", "CO", "MIXTA"):
        pt = "MIXTA"
    with engine.begin() as conn:
        row = conn.execute(text("SELECT id, pais_tipo FROM campanas WHERE nombre=:n"), {"n": nom}).fetchone()
        if row:
            return {"ok": True, "id": int(row[0]), "nombre": nom, "pais_tipo": row[1], "creada": False}
        rid = conn.execute(text("INSERT INTO campanas (nombre, pais_tipo) VALUES (:n, :pt) RETURNING id"),
                           {"n": nom, "pt": pt}).fetchone()[0]
    return {"ok": True, "id": int(rid), "nombre": nom, "pais_tipo": pt, "creada": True}


@app.get("/campana-cargas")
def campana_cargas_get(x_api_key: str = Header(None), campana: str = None, tipo_carga: str = "planeacion"):
    """Devuelve la config de carga de una campaña (para que el frontend elija el cargador)."""
    check_key(x_api_key)
    tipo_carga = (tipo_carga or "planeacion").strip().lower()
    engine = get_engine()
    cid = _id_campana(engine, campana)
    if cid is None:
        raise HTTPException(400, f"La campaña '{campana}' no existe.")
    cfg = _leer_config_carga(engine, cid, tipo_carga)
    if cfg is None:
        # Sin config -> formato clásico (CSV por intervalo de Endesa, endpoint /historico/actualizar)
        return {"ok": True, "tiene_config": False, "formato": "csv_clasico", "tipo_carga": tipo_carga}
    return {"ok": True, "tiene_config": True, "formato": cfg.get("formato"),
            "tipo_carga": tipo_carga, "config": cfg}


@app.post("/campana-cargas")
async def campana_cargas_post(
    x_api_key: str = Header(None),
    campana: str = Form(...),
    tipo_carga: str = Form("planeacion"),
    config: str = Form(...),  # JSON (string)
):
    """Crea/actualiza la config de carga de una campaña (alta sin tocar código)."""
    check_key(x_api_key)
    tipo_carga = (tipo_carga or "planeacion").strip().lower()
    import json
    engine = get_engine()
    cid = _id_campana(engine, campana)
    if cid is None:
        raise HTTPException(400, f"La campaña '{campana}' no existe.")
    try:
        cfg = json.loads(config)
    except Exception as e:
        raise HTTPException(400, f"'config' no es JSON válido: {e}")
    with engine.begin() as conn:
        conn.execute(text("""
            INSERT INTO campana_cargas (campana_id, tipo_carga, config, actualizado_en)
            VALUES (:c, :t, cast(:cfg as jsonb), now())
            ON CONFLICT (campana_id, tipo_carga)
            DO UPDATE SET config = EXCLUDED.config, actualizado_en = now()
        """), {"c": cid, "t": tipo_carga, "cfg": json.dumps(cfg)})
    return {"ok": True, "campana": campana, "tipo_carga": tipo_carga, "config": cfg}


# Sinónimos de columnas para el mapeo de AGENTES (canónico -> posibles nombres en el archivo)
AGENTES_CAMPOS = {
    "dni":                ["dni", "documento", "cedula", "cédula", "nif", "identificacion", "identificación", "id"],
    "nombre":             ["nombre", "nombre completo", "agente", "empleado", "name"],
    "centro":             ["centro", "sede", "site", "ciudad", "location"],
    "pais":               ["pais", "país", "country"],
    "modo":               ["modo", "modalidad"],
    "turno":              ["turno", "shift"],
    "login_acd":          ["login_acd", "login acd", "usuario acd", "login", "usuario", "extension", "extensión", "ext"],
    "fecha_alta":         ["fecha_alta", "fecha alta", "fecha ingreso", "fecha de alta", "ingreso", "alta", "hire date"],
    "jornada_horas":      ["jornada_horas", "jornada", "horas", "horas semana", "hours"],
    "salario_mensual":    ["salario_mensual", "salario", "sueldo", "salary"],
    "vacaciones_anuales": ["vacaciones_anuales", "vacaciones", "dias vacaciones", "días vacaciones", "vacation"],
    "entra_roster":       ["entra_roster", "entra roster", "entra al roster", "roster", "optimiza", "optimizar", "rosterizable"],
}


@app.post("/detectar")
async def detectar(
    x_api_key: str = Header(None),
    archivo: UploadFile = File(...),
    tipo_carga: str = Form("planeacion"),
    hoja: str = Form(None),   # opcional: forzar una hoja concreta
):
    """
    Inspecciona un archivo (Excel/CSV) SIN escribir nada y devuelve lo que el asistente
    de configuración necesita para pintar los desplegables: hojas, columnas, una muestra
    de filas y un mapeo SUGERIDO (fuzzy-match). Pensado para el flujo "cargar y mapear".
    """
    check_key(x_api_key)
    tipo_carga = (tipo_carga or "planeacion").strip().lower()
    file_bytes = await archivo.read()
    nombre = (archivo.filename or "").lower()
    es_csv = nombre.endswith(".csv")

    def _puntua_hoja(cols):
        if tipo_carga == "agentes":
            dni = _col_por_nombre(cols, *AGENTES_CAMPOS["dni"])
            nom = _col_por_nombre(cols, *AGENTES_CAMPOS["nombre"])
            return (2 if dni else 0) + (1 if nom else 0)
        cf = _col_por_nombre(cols, "fecha", "dia", "día", "day")
        cv = _col_por_nombre(cols, "entrantes", "recibidas", "llamadas entrantes",
                             "ofrecidas", "llamadas", "volumen", "vol")
        ci = _col_por_nombre(cols, "intervalo", "hora", "franja", "tramo", "interval")
        cg = _col_por_nombre(cols, "tipo general", "tipo", "cola", "skill", "servicio", "comercializadora")
        return (1 if cf else 0) + (2 if cv else 0) + (1 if ci else 0) + (1 if cg else 0)

    hojas = []
    hoja_activa = None
    if not es_csv:
        try:
            from openpyxl import load_workbook
            wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
            hojas = list(wb.sheetnames)
        except Exception as e:
            raise HTTPException(400, f"No pude abrir el Excel: {type(e).__name__}: {e}")
        if hoja in hojas:
            hoja_activa = hoja
        else:
            # Elegir la hoja que más "pinta de datos" tiene (la que trae fecha + volumen),
            # leyendo solo la cabecera de cada hoja (rápido). Evita agarrar hojas de análisis.
            mejor, mejor_sc = None, -1
            for h in hojas:
                try:
                    cc = [str(c).strip() for c in
                          pd.read_excel(io.BytesIO(file_bytes), sheet_name=h, nrows=0, engine="openpyxl").columns]
                except Exception:
                    cc = []
                sc = _puntua_hoja(cc)
                if sc > mejor_sc:
                    mejor, mejor_sc = h, sc
            hoja_activa = mejor or (hojas[0] if hojas else None)

    try:
        if es_csv:
            df = pd.read_csv(io.BytesIO(file_bytes))
        else:
            df = pd.read_excel(io.BytesIO(file_bytes), sheet_name=hoja_activa, engine="openpyxl")
    except Exception as e:
        raise HTTPException(400, f"No pude leer los datos: {type(e).__name__}: {e}")
    df.columns = [str(c).strip() for c in df.columns]
    columnas = list(df.columns)

    muestra = []
    for _, r in df.head(3).iterrows():
        muestra.append({c: (None if pd.isna(r[c]) else str(r[c])) for c in columnas})

    sugerencia = {}
    if tipo_carga == "planeacion":
        c_fecha = _col_por_nombre(columnas, "fecha", "dia", "día", "day")
        c_int = _col_por_nombre(columnas, "intervalo", "hora", "franja", "tramo", "interval")
        c_vol = _col_por_nombre(columnas, "entrantes", "recibidas", "llamadas entrantes",
                                "ofrecidas", "llamadas", "volumen", "vol")
        c_grupo = _col_por_nombre(columnas, "tipo general", "tipo", "cola", "skill",
                                  "servicio", "comercializadora")
        sugerencia = {"col_fecha": c_fecha, "col_intervalo": c_int,
                      "col_volumen": c_vol, "col_grupo": c_grupo}
        if c_grupo:
            sugerencia["valores_grupo"] = [str(x) for x in df[c_grupo].dropna().unique().tolist()][:50]
        hoja_curva = next((h for h in hojas if "curva" in _norm_txt(h)), None)
        if hoja_curva:
            sugerencia["hoja_curva"] = hoja_curva
            sugerencia["formato_sugerido"] = "diario_curva"
        elif c_int:
            sugerencia["formato_sugerido"] = "intervalo"

    elif tipo_carga == "agentes":
        cols_map = {campo: _col_por_nombre(columnas, *syns) for campo, syns in AGENTES_CAMPOS.items()}
        sugerencia = {"columnas": cols_map}

    elif tipo_carga == "turnos":
        if es_csv:
            sugerencia = {"aviso": "La parrilla de turnos debe ser un Excel (.xlsx/.xlsm)."}
        else:
            from openpyxl import load_workbook
            wb2 = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
            sn = hoja_activa if hoja_activa in wb2.sheetnames else (wb2.sheetnames[0] if wb2.sheetnames else None)
            rows = list(wb2[sn].iter_rows(values_only=True)) if sn else []
            fila_cab, col_dni, cols_fecha = _detecta_layout_parrilla(rows)
            sugerencia = {
                "fila_cabecera": fila_cab,
                "col_dni": col_dni,
                "n_fechas": len(cols_fecha),
                "codigos_detectados": _codigos_de_parrilla(rows, fila_cab, col_dni, cols_fecha),
            }

    elif tipo_carga == "conexiones":
        df2, fila_cab = _df_con_cabecera_auto(file_bytes, archivo.filename, hoja_activa)
        columnas = list(df2.columns)
        muestra = []
        for _, rr in df2.head(3).iterrows():
            muestra.append({c: (None if pd.isna(rr[c]) else str(rr[c])) for c in columnas})
        cmap = {campo: _buscar_col(columnas, syns) for campo, syns in CONEXIONES_CAMPOS.items()}
        sugerencia = {"fila_cabecera": fila_cab, "columnas": cmap}

    return {
        "ok": True,
        "tipo_carga": tipo_carga,
        "es_csv": es_csv,
        "hojas": hojas,
        "hoja_activa": hoja_activa,
        "columnas": columnas,
        "muestra": muestra,
        "sugerencia": sugerencia,
    }


@app.post("/historico/importar")
async def historico_importar(
    x_api_key: str = Header(None),
    archivo: UploadFile = File(...),
    campana: str = Form(...),
    tipo_carga: str = Form("planeacion"),
    reemplazar: bool = Form(False),   # True: borra antes el histórico de las colas tocadas
):
    """
    Carga GENÉRICA de histórico de planeación, guiada por la config de la campaña
    (tabla campana_cargas, tipo_carga='planeacion'). Soporta dos formatos:

      • formato 'intervalo'  -> el archivo ya trae datos por intervalo (granularidad
        60/30/15 min, campo 'granularidad_min'). Se AGREGAN a HORA (0-23) y se escriben.
      • formato 'diario_curva' -> volumen DIARIO + hoja de curva intradía; cada día se
        EXPANDE a 24 horas con la curva, preservando el total diario exacto.

    Las 'colas' se definen en la config (1..N), cada una con su 'incluye' (lista de
    valores de la columna de grupo, o null = todo). Escribe en historico_llamadas a
    granularidad HORARIA (el motor dimensiona por hora). Sin código específico de campaña.
    """
    check_key(x_api_key)
    tipo_carga = (tipo_carga or "planeacion").strip().lower()
    engine = get_engine()
    cid = _id_campana(engine, campana)
    if cid is None:
        raise HTTPException(400, f"La campaña '{campana}' no existe.")
    cfg = _leer_config_carga(engine, cid, tipo_carga)
    if cfg is None:
        raise HTTPException(400, f"La campaña '{campana}' no tiene config de carga '{tipo_carga}'. "
                                 f"Créala en /campana-cargas o usa el cargador CSV clásico.")

    formato = str(cfg.get("formato", "")).lower()
    col_fecha_cfg = cfg.get("col_fecha", "fecha")
    col_vol_cfg = cfg.get("col_volumen", "entrantes")
    col_grupo_cfg = cfg.get("col_grupo")
    colas_cfg = cfg.get("colas") or [{"nombre": "Total", "incluye": None}]
    hoja_datos = cfg.get("hoja_datos")
    gran = int(cfg.get("granularidad_min", 60) or 60)

    file_bytes = await archivo.read()
    try:
        df = _leer_tabla(file_bytes, archivo.filename, hoja_datos)
    except Exception as e:
        raise HTTPException(400, f"No pude leer los datos (hoja '{hoja_datos}'): {type(e).__name__}: {e}")
    df.columns = [str(c).strip() for c in df.columns]

    c_fecha = _col_por_nombre(df.columns, col_fecha_cfg)
    c_vol = _col_por_nombre(df.columns, col_vol_cfg, "entrantes", "llamadas", "volumen")
    c_grupo = _col_por_nombre(df.columns, col_grupo_cfg) if col_grupo_cfg else None
    if not c_fecha or not c_vol:
        raise HTTPException(400, f"No encuentro columnas fecha='{col_fecha_cfg}' / volumen='{col_vol_cfg}'. "
                                 f"Columnas vistas: {list(df.columns)}")
    df[c_fecha] = pd.to_datetime(df[c_fecha], errors="coerce")
    df = df.dropna(subset=[c_fecha])
    df[c_vol] = pd.to_numeric(df[c_vol], errors="coerce").fillna(0)
    total_bruto = int(df[c_vol].sum())

    filas = []          # {f, h, c(cola), e}
    resumen_colas = []  # por cola: dias, entrantes, filas_hora

    if formato == "diario_curva":
        curva = _curva_intradia_desde_excel(file_bytes, cfg.get("hoja_curva", "Curvas"))
        for cola in colas_cfg:
            nombre = cola["nombre"]
            sub = _filtra_por_grupo(df, c_grupo, cola.get("incluye"))
            diario = {f: int(round(v)) for f, v in sub.groupby(sub[c_fecha].dt.date)[c_vol].sum().items()}
            nf = tot = 0
            for f, total in diario.items():
                pesos = curva.get(f.weekday(), [1 / 24] * 24)
                s = sum(pesos)
                pesos = [p / s for p in pesos] if s > 0 else [1 / 24] * 24
                ph = _reparte_entero(total, pesos)
                for h in range(24):
                    if ph[h] > 0:
                        filas.append({"f": f, "h": h, "c": nombre, "e": ph[h]}); nf += 1; tot += ph[h]
            resumen_colas.append({"cola": nombre, "dias": len(diario), "entrantes": tot, "filas_hora": nf})

    elif formato == "intervalo":
        c_int = _col_por_nombre(df.columns, cfg.get("col_intervalo", "hora"), "hora", "intervalo")
        if not c_int:
            raise HTTPException(400, f"Formato 'intervalo' requiere columna de intervalo "
                                     f"'{cfg.get('col_intervalo', 'hora')}'. Columnas: {list(df.columns)}")
        df["_hora"] = df[c_int].map(lambda v: _a_hora_de_intervalo(v, gran))
        df = df.dropna(subset=["_hora"])
        df["_hora"] = df["_hora"].astype(int)
        for cola in colas_cfg:
            nombre = cola["nombre"]
            sub = _filtra_por_grupo(df, c_grupo, cola.get("incluye"))
            agg = sub.groupby([sub[c_fecha].dt.date, "_hora"])[c_vol].sum()
            dias = set(); nf = tot = 0
            for (f, h), v in agg.items():
                e = int(round(v))
                if e > 0:
                    filas.append({"f": f, "h": int(h), "c": nombre, "e": e})
                    nf += 1; tot += e; dias.add(f)
            resumen_colas.append({"cola": nombre, "dias": len(dias), "entrantes": tot, "filas_hora": nf})
    else:
        raise HTTPException(400, f"Formato de carga no soportado: '{formato}'. Usa 'intervalo' o 'diario_curva'.")

    if not filas:
        raise HTTPException(400, "No se generaron filas (revisa columnas, colas/incluye y datos).")

    nombres_colas = sorted({fl["c"] for fl in filas})
    insertadas = 0
    with engine.begin() as conn:
        if reemplazar:
            inc, pin = _in_clause("k", nombres_colas)
            conn.execute(text(f"DELETE FROM historico_llamadas WHERE campana_id=:c AND cola IN {inc}"),
                         {"c": cid, **pin})
        LOTE = 500
        for k in range(0, len(filas), LOTE):
            lote = filas[k:k + LOTE]
            vals = []
            params = {"cid": cid}
            for j, r in enumerate(lote):
                vals.append(f"(:f{j}, :h{j}, :c{j}, :e{j}, :e{j}, 0, :cid, now())")
                params[f"f{j}"] = r["f"]
                params[f"h{j}"] = r["h"]
                params[f"c{j}"] = r["c"]
                params[f"e{j}"] = r["e"]
            sql = ("INSERT INTO historico_llamadas "
                   "(fecha, hora, cola, entrantes, atendidas, abandonadas, campana_id, actualizado_en) "
                   "VALUES " + ", ".join(vals) +
                   " ON CONFLICT (fecha, hora, cola, campana_id) DO UPDATE SET "
                   "entrantes=EXCLUDED.entrantes, atendidas=EXCLUDED.atendidas, "
                   "abandonadas=EXCLUDED.abandonadas, actualizado_en=now()")
            conn.execute(text(sql), params)
            insertadas += len(lote)

    inc, pin = _in_clause("k", nombres_colas)
    rango = pd.read_sql(text("SELECT MIN(fecha) d, MAX(fecha) h, COUNT(*) n, COALESCE(SUM(entrantes),0) e "
                             f"FROM historico_llamadas WHERE campana_id=:c AND cola IN {inc}"),
                        engine, params={"c": cid, **pin})
    return {
        "ok": True,
        "campana": campana,
        "formato": formato,
        "granularidad_min": gran,
        "total_entrantes_bruto": total_bruto,
        "colas": resumen_colas,
        "filas_hora_insertadas": insertadas,
        "historico_desde": str(rango["d"][0]),
        "historico_hasta": str(rango["h"][0]),
        "total_filas": int(rango["n"][0]),
        "total_entrantes": int(rango["e"][0]),
    }


@app.post("/planeacion")
async def planeacion(
    x_api_key: str = Header(None),
    mes: str = Form(...),
    archivo: UploadFile = File(None),
    aht: int = Form(420),
    sla: float = Form(0.80),
    asa: int = Form(20),
    occ: float = Form(0.65),
    utl: float = Form(0.88),
    esp_max: int = Form(12),
    largo: int = Form(9),
    nda_obj: float = Form(0.96),
    paciencia: int = Form(90),
    estructura: str = Form("mixto"),
    absentismo: float = Form(0.15),
    campana: str = Form("Endesa"),
    ajuste_pct: float = Form(0.0),   # % de ajuste de volumen (ej. 10 = +10%, -5 = -5%)
    comparar_skills: bool = Form(False),  # True: añade comparación dedicado vs multiskill por cola
):
    """Corre el optimizador y devuelve datos para graficar SIN escribir en la base."""
    check_key(x_api_key)
    if archivo is not None:
        file_bytes = await archivo.read()
    else:
        eng = get_engine()
        file_bytes = _historico_a_csv_bytes(eng, _id_campana(eng, campana))
    ajuste = (ajuste_pct or 0.0) / 100.0
    largo_df = motor.largo_desde_historico(file_bytes, mes, "Nacional España", 4, 6, ajuste, mixto=False)
    S = motor.dimension_roster(largo_df, aht, sla, asa, occ, utl, esp_max, largo,
                               nda_obj, paciencia, estructura)
    turnos = motor.turnos_dict(S, largo)

    # Datos por día (0=Lun .. 6=Dom): requerido, programado, ocupación, NDA
    NOM = ["Lunes", "Martes", "Miércoles", "Jueves", "Viernes", "Sábado", "Domingo"]
    por_dia = []
    for d in range(7):
        req = [S["peak"][d][h] for h in range(24)]
        cob = [motor.cubierto(S, turnos, d, h) for h in range(24)]
        occv = [round(S["occ"][d][h] * 100, 1) for h in range(24)]
        ndav = [round(S["nda"][d][h] * 100, 1) for h in range(24)]
        por_dia.append({"dia": d, "nombre": NOM[d], "requerido": req,
                        "programado": cob, "ocupacion": occv, "nda": ndav})

    # Plan de turnos (tabla)
    plan = []
    for k in sorted(S["xe"], key=int):
        t = int(k)
        plan.append({"pais": "España", "inicio": f"{t:02d}:00",
                     "fin": f"{(t+largo)%24:02d}:00", "cantidad": S["xe"][k],
                     "libres": "Sáb, Dom"})
    for k in sorted(S["xc"], key=lambda x: (int(x.split("_")[0]), int(x.split("_")[1]))):
        t, p = map(int, k.split("_"))
        o = sorted(motor.LIBRES_VIZ[p])
        plan.append({"pais": "Colombia", "inicio": f"{t:02d}:00",
                     "fin": f"{(t+largo)%24:02d}:00", "cantidad": S["xc"][k],
                     "libres": f"{NOM[o[0]][:3]}, {NOM[o[1]][:3]}"})

    import math as _m
    # Headline POOLED (multiskill) por defecto = comportamiento actual
    presentes = S["te"] + S["tc"]
    en_nomina = _m.ceil(presentes / (1 - absentismo)) if absentismo < 1 else presentes
    head_espana = S["te"]; head_colombia = S["tc"]

    # --- Fase 2: modo de dimensionamiento por campaña (dedicado | multiskill) ---
    # Vive en la config de planeación: {"skills": {"modo": "dedicado"}}. Default multiskill.
    cid_cmp = _id_campana(get_engine(), campana)
    cfg_cmp = _leer_config_carga(get_engine(), cid_cmp, "planeacion") if cid_cmp else None
    modo_dim = (((cfg_cmp or {}).get("skills") or {}).get("modo") or "multiskill")

    # Dimensionado DEDICADO (cola por cola). Se calcula si el usuario pide comparar,
    # o si la campaña está en modo 'dedicado' (para usarlo como headline). N corridas
    # del motor (1 por cola); validado en local. No rompe el cálculo si falla.
    comparacion_skills = None
    dedicado_total = None
    if comparar_skills or modo_dim == "dedicado":
        try:
            nombres = [c.get("nombre") for c in ((cfg_cmp or {}).get("colas") or []) if c.get("nombre")]
            dedicado = []
            for nom in nombres:
                ldf = motor.largo_desde_historico(file_bytes, mes, "Nacional España", 4, 6,
                                                  ajuste, mixto=False, colas_incluidas=[nom])
                Sd = motor.dimension_roster(ldf, aht, sla, asa, occ, utl, esp_max, largo,
                                            nda_obj, paciencia, estructura)
                pres_d = Sd["te"] + Sd["tc"]
                dedicado.append({
                    "cola": nom, "volumen": int(Sd["total"]),
                    "espana_lv": int(Sd["te"]), "colombia_247": int(Sd["tc"]),
                    "presentes": int(pres_d),
                    "en_nomina": _m.ceil(pres_d / (1 - absentismo)) if absentismo < 1 else int(pres_d),
                })
            tot_ded = sum(d["presentes"] for d in dedicado)
            dedicado_total = {
                "presentes": tot_ded,
                "espana": sum(d["espana_lv"] for d in dedicado),
                "colombia": sum(d["colombia_247"] for d in dedicado),
            }
            if comparar_skills:
                _pool = S["te"] + S["tc"]
                comparacion_skills = {
                    "dedicado_por_cola": dedicado,
                    "total_dedicado_presentes": tot_ded,
                    "multiskill_presentes": _pool,
                    "ahorro_pooling": tot_ded - _pool,
                    "ahorro_pct": round(100.0 * (tot_ded - _pool) / tot_ded, 1) if tot_ded else 0,
                }
        except Exception as e:
            comparacion_skills = {"error": f"{type(e).__name__}: {str(e)[:200]}"}

    # Si la campaña es 'dedicado', el headline pasa a ser el total dedicado
    if modo_dim == "dedicado" and dedicado_total:
        presentes = dedicado_total["presentes"]
        en_nomina = _m.ceil(presentes / (1 - absentismo)) if absentismo < 1 else presentes
        head_espana = dedicado_total["espana"]; head_colombia = dedicado_total["colombia"]

    # Guardar los parámetros usados + resumen del resultado para esta campaña+mes
    cid_save = _id_campana(get_engine(), campana)
    if cid_save is not None:
        try:
            with get_engine().begin() as conn:
                conn.execute(text("""
                    insert into campana_parametros
                      (campana_id, mes, aht, sla, asa, occ, utl, esp_max, largo, nda_obj,
                       paciencia, absentismo, estructura,
                       res_volumen, res_espana, res_colombia, res_presentes, res_en_nomina,
                       calculado_en, actualizado_en)
                    values
                      (:cid, :mes, :aht, :sla, :asa, :occ, :utl, :esp_max, :largo, :nda_obj,
                       :paciencia, :absentismo, :estructura,
                       :rv, :re, :rc, :rp, :rn, now(), now())
                    on conflict (campana_id, mes) do update set
                      aht=excluded.aht, sla=excluded.sla, asa=excluded.asa, occ=excluded.occ,
                      utl=excluded.utl, esp_max=excluded.esp_max, largo=excluded.largo,
                      nda_obj=excluded.nda_obj, paciencia=excluded.paciencia,
                      absentismo=excluded.absentismo, estructura=excluded.estructura,
                      res_volumen=excluded.res_volumen, res_espana=excluded.res_espana,
                      res_colombia=excluded.res_colombia, res_presentes=excluded.res_presentes,
                      res_en_nomina=excluded.res_en_nomina,
                      calculado_en=now(), actualizado_en=now()
                """), {"cid": cid_save, "mes": mes, "aht": aht, "sla": sla, "asa": asa,
                       "occ": occ, "utl": utl, "esp_max": esp_max, "largo": largo,
                       "nda_obj": nda_obj, "paciencia": paciencia, "absentismo": absentismo,
                       "estructura": estructura, "rv": int(S["total"]), "re": int(head_espana),
                       "rc": int(head_colombia), "rp": int(presentes), "rn": int(en_nomina)})
        except Exception:
            pass  # si falla el guardado del resumen, no rompemos el cálculo

    return {
        "ok": True,
        "mes": mes,
        "volumen_total": S["total"],
        "estructura": estructura,
        "plantilla": {
            "espana_lv": head_espana,
            "colombia_247": head_colombia,
            "total_presentes": presentes,
            "en_nomina": en_nomina,
        },
        "modo_dimensionamiento": modo_dim,
        "por_dia": por_dia,
        "plan_turnos": plan,
        "objetivos": {"occ": round(occ * 100, 1), "nda": round(nda_obj * 100, 1)},
        "comparacion_skills": comparacion_skills,
    }


@app.post("/generar-roster/huerfanos")
async def roster_huerfanos(
    x_api_key: str = Header(None),
    mes: str = Form(...),            # "2026-06"
    campana: str = Form(...),
):
    """Detecta 'huérfanos': agentes que ya NO entran al roster (entra_roster=false)
    pero tienen turnos del optimizador en el mes. Solo LEE; no modifica nada.
    Sirve para que el WFM decida conservar/liberar ANTES de regenerar (paso 1 de 2)."""
    check_key(x_api_key)
    engine = get_engine()
    cid = _id_campana(engine, campana)
    if cid is None:
        raise HTTPException(400, f"La campaña '{campana}' no existe.")
    dias = pd.date_range(f"{mes}-01", pd.Timestamp(f"{mes}-01") + pd.offsets.MonthEnd(0))
    df = pd.read_sql(text("""
        SELECT a.id AS agente_id, a.nombre, a.centro, a.pais,
               SUM(CASE WHEN s.tipo='trabajo' THEN 1 ELSE 0 END) AS dias_trabajo,
               COUNT(*) AS dias_total,
               MAX(CASE WHEN s.bloqueado IS TRUE THEN 1 ELSE 0 END) AS tiene_bloqueados
        FROM asignaciones s JOIN agentes a ON a.id = s.agente_id
        WHERE s.campana_id = :c AND s.fecha BETWEEN :i AND :f
          AND a.entra_roster = false
          AND s.creado_por = 'optimizador-web'
        GROUP BY a.id, a.nombre, a.centro, a.pais
        ORDER BY a.nombre
    """), engine, params={"c": int(cid), "i": dias[0].date(), "f": dias[-1].date()})
    huerfanos = []
    for _, r in df.iterrows():
        huerfanos.append({
            "agente_id": int(r["agente_id"]), "nombre": r["nombre"],
            "centro": r["centro"], "pais": r["pais"],
            "dias_trabajo": int(r["dias_trabajo"] or 0),
            "dias_total": int(r["dias_total"] or 0),
            "tiene_bloqueados": bool(r["tiene_bloqueados"]),
        })
    return {"ok": True, "mes": mes, "campana": campana,
            "total_huerfanos": len(huerfanos), "huerfanos": huerfanos}


@app.post("/generar-roster")
async def generar_roster(
    x_api_key: str = Header(None),
    mes: str = Form(...),                 # "2026-07"
    campana: str = Form("Endesa"),
    archivo: UploadFile = File(None),     # historico.csv (opcional: si no, usa histórico de base)
    aht: int = Form(420),
    sla: float = Form(0.80),
    asa: int = Form(20),
    occ: float = Form(0.65),
    utl: float = Form(0.88),
    esp_max: int = Form(12),
    largo: int = Form(9),
    nda_obj: float = Form(0.96),
    paciencia: int = Form(90),
    estructura: str = Form("mixto"),
    ajuste_pct: float = Form(0.0),
    conservar_ids: str = Form(""),   # Pieza E: ids de huérfanos a CONSERVAR (se bloquean)
    liberar_ids: str = Form(""),     # Pieza E: ids de huérfanos a LIBERAR (se borran)
):
    check_key(x_api_key)
    engine = get_engine()

    # 1) Leer histórico (del CSV subido o del histórico acumulado en la base) y correr optimizador
    if archivo is not None:
        file_bytes = await archivo.read()
    else:
        file_bytes = _historico_a_csv_bytes(engine, _id_campana(engine, campana))
    ajuste = (ajuste_pct or 0.0) / 100.0
    largo_df = motor.largo_desde_historico(file_bytes, mes, "Nacional España", 4, 6, ajuste, mixto=False)
    S = motor.dimension_roster(largo_df, aht, sla, asa, occ, utl, esp_max, largo,
                               nda_obj, paciencia, estructura)

    # 2) Repartir personas
    id_campana = pd.read_sql(text("SELECT id FROM campanas WHERE nombre=:n"),
                             engine, params={"n": campana})["id"][0]

    dias_mes = pd.date_range(f"{mes}-01", pd.Timestamp(f"{mes}-01") + pd.offsets.MonthEnd(0))

    # --- BLOQUEOS: asignaciones que el WFM marcó como fijas y NO deben recalcularse ---
    bloq = pd.read_sql(text("""
        SELECT agente_id, fecha FROM asignaciones
        WHERE campana_id=:c AND fecha BETWEEN :i AND :f AND bloqueado = true
    """), engine, params={"c": int(id_campana), "i": dias_mes[0].date(), "f": dias_mes[-1].date()})
    # set de (agente_id, fecha) bloqueados — esos días no se regeneran
    dias_bloqueados = set((int(r["agente_id"]), pd.Timestamp(r["fecha"]).date()) for _, r in bloq.iterrows())
    # agentes con TODO el mes bloqueado salen del reparto automático
    cont_bloq = bloq.groupby("agente_id").size() if not bloq.empty else pd.Series(dtype=int)
    agentes_full_bloq = set(int(a) for a, n in cont_bloq.items() if n >= len(dias_mes))

    # Reciben turno del optimizador los agentes de ESTA campaña marcados como
    # 'entra_roster' (los que no entran -backoffice, formación- se asignan aparte).
    # Antes filtraba por modo='MULTISKILL' (overload del campo) y SIN campana_id
    # (arrastraba agentes de otras campañas). Ahora: por campaña + entra_roster.
    ag = pd.read_sql(text("SELECT id,nombre,centro,pais,jornada_horas,modo FROM agentes "
                          "WHERE estado='ACTIVO' AND entra_roster = true AND campana_id = :cid "
                          "ORDER BY id"), engine, params={"cid": int(id_campana)})
    ag = ag[~ag["id"].isin(agentes_full_bloq)]  # excluir agentes totalmente bloqueados
    esp = ag[ag["pais"] == "España"].to_dict("records")
    col = ag[ag["pais"] == "Colombia"].to_dict("records")
    # Perfil de tráfico por hora (turnos que pidió el optimizador) — define dónde arrancan más agentes
    col_t, esp_t = repartidor.demanda_diaria(S)
    # Repartir TODOS los Multiskill (sin backup), según el perfil de tráfico
    base_col, res_col = repartidor.repartir_todos(col, col_t, set(range(7)), "Colombia", True)
    base_esp, res_esp = repartidor.repartir_todos(esp, esp_t, {0, 1, 2, 3, 4}, "España", False)
    base_todos = base_col + base_esp

    # 3) Registrar turnos usados
    horas_usadas = sorted({a["hora_inicio"] for a in base_todos})
    turno_id_por_hora = {}
    with engine.begin() as conn:
        for h in horas_usadas:
            hora_str = f"{h:02d}:00"
            row = conn.execute(text("SELECT id FROM turnos WHERE hora_inicio=:hi AND duracion_horas=:du LIMIT 1"),
                               {"hi": hora_str, "du": largo}).fetchone()
            if row:
                turno_id_por_hora[h] = row[0]
            else:
                r = conn.execute(text("INSERT INTO turnos (nombre,hora_inicio,duracion_horas,descripcion) VALUES (:n,:hi,:du,:de) RETURNING id"),
                                 {"n": f"Turno {hora_str} ({largo}h)", "hi": hora_str, "du": largo, "de": f"Optimizador {mes}"}).fetchone()
                turno_id_por_hora[h] = r[0]

    # 4) Generar asignaciones (saltando los días bloqueados por el WFM)
    filas = []
    for item in base_todos:
        a = item["agente"]; hi = item["hora_inicio"]; libres = item["libres"]
        hora_fin = (hi + largo) % 24; tid = turno_id_por_hora[hi]
        for d in dias_mes:
            # si ese día de ese agente está bloqueado, no lo regeneramos (se conserva el existente)
            if (int(a["id"]), d.date()) in dias_bloqueados:
                continue
            if d.dayofweek in libres:
                filas.append({"agente_id": int(a["id"]), "fecha": d.date(), "campana_id": int(id_campana),
                              "turno_id": None, "hora_inicio": None, "hora_fin": None, "tipo": "libre", "creado_por": "optimizador-web"})
            else:
                filas.append({"agente_id": int(a["id"]), "fecha": d.date(), "campana_id": int(id_campana),
                              "turno_id": int(tid), "hora_inicio": f"{hi:02d}:00", "hora_fin": f"{hora_fin:02d}:00", "tipo": "trabajo", "creado_por": "optimizador-web"})
    n_asig = len(filas)

    # --- Pieza E: gestión de huérfanos (agentes fuera del roster con turnos previos) ---
    #   conservar_ids -> se BLOQUEAN sus turnos del optimizador (sobreviven al regenerado).
    #   liberar_ids   -> se BORRAN sus turnos del optimizador (incluso si estaban bloqueados).
    # Se procesa ANTES del DELETE/insert para que el conservar quede protegido.
    def _ids_lista(s):
        out = []
        for x in str(s or "").split(","):
            x = x.strip()
            if x:
                try: out.append(int(x))
                except ValueError: pass
        return out
    _cons = _ids_lista(conservar_ids); _libr = _ids_lista(liberar_ids)
    if _cons or _libr:
        with engine.begin() as conn:
            if _cons:
                _in, _p = _in_clause("ag", _cons)
                _p.update({"c": int(id_campana), "i": dias_mes[0].date(), "f": dias_mes[-1].date()})
                conn.execute(text(f"UPDATE asignaciones SET bloqueado = true "
                                  f"WHERE campana_id=:c AND fecha BETWEEN :i AND :f "
                                  f"AND creado_por='optimizador-web' AND agente_id IN {_in}"), _p)
            if _libr:
                _in, _p = _in_clause("ag", _libr)
                _p.update({"c": int(id_campana), "i": dias_mes[0].date(), "f": dias_mes[-1].date()})
                conn.execute(text(f"DELETE FROM asignaciones "
                                  f"WHERE campana_id=:c AND fecha BETWEEN :i AND :f "
                                  f"AND creado_por='optimizador-web' AND agente_id IN {_in}"), _p)

    # 5) Cargar asignaciones (limpiar mes antes, PERO conservando las bloqueadas)
    #    INSERT MULTI-FILA con ON CONFLICT: mete muchas filas por sentencia (rápido),
    #    y si una (agente_id, fecha) ya existe, la actualiza SOLO si no está bloqueada.
    #    Así nunca choca con la restricción única y respeta los turnos bloqueados por el WFM.
    try:
        # DELETE de las no-bloqueadas
        with engine.begin() as conn:
            conn.execute(text("DELETE FROM asignaciones WHERE campana_id=:c AND fecha>=:i AND fecha<=:f AND (bloqueado IS NULL OR bloqueado = false)"),
                         {"c": int(id_campana), "i": dias_mes[0].date(), "f": dias_mes[-1].date()})

        cols = ["agente_id", "fecha", "campana_id", "turno_id", "hora_inicio", "hora_fin", "tipo", "creado_por"]
        conflict = """
            ON CONFLICT (agente_id, fecha) DO UPDATE SET
                campana_id = EXCLUDED.campana_id,
                turno_id = EXCLUDED.turno_id,
                hora_inicio = EXCLUDED.hora_inicio,
                hora_fin = EXCLUDED.hora_fin,
                tipo = EXCLUDED.tipo,
                creado_por = EXCLUDED.creado_por
            WHERE asignaciones.bloqueado IS NOT TRUE
        """
        LOTE = 500
        with engine.begin() as conn:
            for k in range(0, len(filas), LOTE):
                lote = filas[k:k+LOTE]
                valores = []
                params = {}
                for j, f in enumerate(lote):
                    valores.append(f"(:agente_id{j}, :fecha{j}, :campana_id{j}, :turno_id{j}, :hora_inicio{j}, :hora_fin{j}, :tipo{j}, :creado_por{j})")
                    for col in cols:
                        params[f"{col}{j}"] = f[col]
                sql = "INSERT INTO asignaciones (" + ", ".join(cols) + ") VALUES " + ", ".join(valores) + conflict
                conn.execute(text(sql), params)
    except Exception as e:
        ejemplo = filas[0] if filas else {}
        raise HTTPException(500, f"Error al escribir asignaciones: {type(e).__name__}: {str(e)[:300]} | ejemplo fila: {ejemplo}")

    # 6) Los breaks se generan en un endpoint aparte (/generar-breaks) para que esta
    #    respuesta sea rápida y no exceda el límite de tiempo de Cloudflare/Lovable.

    return {
        "ok": True,
        "mes": mes,
        "volumen_total": S["total"],
        "asignaciones": n_asig,
        "breaks": "pendiente",
        "nota": "Asignaciones generadas. Los breaks se generan en el paso siguiente.",
        "colombia": res_col,
        "espana": res_esp,
    }


@app.post("/generar-breaks")
def generar_breaks(
    x_api_key: str = Header(None),
    mes: str = Form(...),
    campana: str = Form("Endesa"),
):
    """Genera los breaks del mes a partir de las asignaciones ya creadas.
    Se llama DESPUÉS de /generar-roster para repartir el trabajo y no exceder el tiempo límite."""
    check_key(x_api_key)
    engine = get_engine()
    id_campana = pd.read_sql(text("SELECT id FROM campanas WHERE nombre=:n"),
                             engine, params={"n": campana})["id"][0]
    dias_mes = pd.date_range(f"{mes}-01", pd.Timestamp(f"{mes}-01") + pd.offsets.MonthEnd(0))

    asg = pd.read_sql(text("""
        SELECT a.id AS asignacion_id, a.hora_inicio, a.hora_fin, ag.pais
        FROM asignaciones a JOIN agentes ag ON ag.id=a.agente_id
        WHERE a.campana_id=:c AND a.tipo='trabajo' AND a.fecha BETWEEN :i AND :f
        ORDER BY a.fecha, a.hora_inicio, a.id
    """), engine, params={"c": int(id_campana), "i": dias_mes[0].date(), "f": dias_mes[-1].date()})

    filas_b = []; contador = defaultdict(int)
    for _, r in asg.iterrows():
        dur = dur_turno(r["hora_inicio"], r["hora_fin"]); idx = contador[r["hora_inicio"]]; contador[r["hora_inicio"]] += 1
        gen = breaks_colombia if r["pais"] == "Colombia" else breaks_espana
        for bhi, dm, tp in gen(r["hora_inicio"], round(dur), idx):
            filas_b.append({"asignacion_id": int(r["asignacion_id"]),
                            "hora_inicio": str(bhi), "duracion_min": int(dm), "tipo": str(tp)})

    ids = [int(x) for x in asg["asignacion_id"].tolist()]
    try:
        # DELETE de los breaks viejos: un solo DELETE por todos los ids (con índice es rápido)
        with engine.begin() as conn:
            if ids:
                conn.execute(text("DELETE FROM breaks WHERE asignacion_id = ANY(:ids)"), {"ids": ids})

        # INSERT MULTI-FILA: muchas filas por sentencia (VALUES (...),(...),...),
        # en vez de una fila por viaje. Reduce miles de viajes a la base a unas pocas decenas.
        LOTE = 500
        with engine.begin() as conn:
            for k in range(0, len(filas_b), LOTE):
                lote = filas_b[k:k+LOTE]
                valores = []
                params = {}
                for j, f in enumerate(lote):
                    valores.append(f"(:a{j}, :h{j}, :d{j}, :t{j})")
                    params[f"a{j}"] = f["asignacion_id"]
                    params[f"h{j}"] = f["hora_inicio"]
                    params[f"d{j}"] = f["duracion_min"]
                    params[f"t{j}"] = f["tipo"]
                sql = "INSERT INTO breaks (asignacion_id, hora_inicio, duracion_min, tipo) VALUES " + ", ".join(valores)
                conn.execute(text(sql), params)
    except Exception as e:
        ejemplo = filas_b[0] if filas_b else {}
        raise HTTPException(500, f"Error al escribir breaks: {type(e).__name__}: {str(e)[:300]} | ejemplo: {ejemplo}")

    return {"ok": True, "mes": mes, "breaks": len(filas_b)}


# ============================================================
#  ADHERENCIA · Fase 2 — Importar resumen diario del ACD
# ============================================================
# Mapa: clave logica -> posibles inicios del nombre de columna en el Excel.
# Se busca por "empieza por" para tolerar nombres recortados (ej "Total Not Ready Ti").
_ACD_COLS = {
    "fecha":            ["Fecha"],
    "plataforma":       ["Plataforma"],
    "login_acd":        ["Agente"],
    "logado":           ["Logado"],
    "deslogado":        ["Deslogado"],
    "seg_login":        ["Total Login"],
    "seg_not_ready":    ["Total Not Ready"],
    # Formato nuevo del ACD: Status 1..8 (2=PAUSA,3=DESCANSO,4=FORMACION,5=BO,6=SERVICIO).
    # Si no estan, cae a los nombres del formato viejo.
    "seg_pausa_visual": ["Total Status 2", "Pausa Visual"],
    "seg_break":        ["Total Status 3", "Break"],
    "seg_formacion":    ["Total Status 4", "Formacion", "Formaci\u00f3n"],
    "seg_backoffice":   ["Total Status 5", "BackOffice", "Back Office"],   # BO = productivo
    "seg_servicios":    ["Total Status 6", "Servicios"],
    "seg_talk_in":      ["Total Talk Time In"],
    "seg_talk_out":     ["Total Talk Time Ou"],
    "seg_hold":         ["Total Hold"],
    "llamadas_inbound": ["Total Calls Inboun"],
}


def _buscar_col(columnas, prefijos):
    """Nombre real de la columna que empieza por alguno de los prefijos (ignora may/espacios)."""
    norm = {c: str(c).strip().lower() for c in columnas}
    for p in prefijos:
        pl = p.strip().lower()
        for real, n in norm.items():
            if n.startswith(pl):
                return real
    return None


def _a_int(v):
    try:
        if v is None or (isinstance(v, float) and pd.isna(v)):
            return 0
        return int(round(float(v)))
    except Exception:
        return 0


def _a_txt(v):
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return None
    s = str(v).strip()
    return s if s and s.lower() != "nan" else None


def _a_hora(v):
    """'10:15' o time de Excel -> 'HH:MM:SS' (texto). Vacio -> None."""
    s = _a_txt(v)
    if not s:
        return None
    partes = s.split(":")
    try:
        h = int(partes[0]); m = int(partes[1]) if len(partes) > 1 else 0
        return f"{h:02d}:{m:02d}:00"
    except Exception:
        return None


# Sinónimos de columnas del ACD/conexiones (canónico -> prefijos para _buscar_col)
CONEXIONES_CAMPOS = {
    "fecha":            ["fecha", "date", "dia", "día"],
    "plataforma":       ["plataforma", "platform"],
    "login_acd":        ["agente", "usuario", "login_acd", "login acd", "extension", "extensión", "ext", "id agente"],
    "logado":           ["logado", "logon", "login time", "hora login", "conexion", "conexión"],
    "deslogado":        ["deslogado", "logoff", "logout", "desconexion", "desconexión"],
    "seg_login":        ["total login", "tiempo login", "seg login"],
    "seg_not_ready":    ["total not ready", "not ready", "no disponible", "aux"],
    "seg_pausa_visual": ["total status 2", "pausa visual", "pausa"],
    "seg_break":        ["total status 3", "break", "descanso"],
    "seg_formacion":    ["total status 4", "formacion", "formación", "training"],
    "seg_backoffice":   ["total status 5", "backoffice", "back office"],
    "seg_servicios":    ["total status 6", "servicios"],
    "seg_talk_in":      ["total talk time in", "talk in", "habla entrante"],
    "seg_talk_out":     ["total talk time ou", "talk out", "habla saliente"],
    "seg_hold":         ["total hold", "hold", "espera"],
    "llamadas_inbound": ["total calls inboun", "calls inbound", "inbound", "llamadas"],
}

# Orden de campos numéricos (segundos) del ACD
_ACD_SEGUNDOS = ["seg_login", "seg_not_ready", "seg_pausa_visual", "seg_break", "seg_formacion",
                 "seg_servicios", "seg_backoffice", "seg_talk_in", "seg_talk_out", "seg_hold",
                 "llamadas_inbound"]


def _df_con_cabecera_auto(file_bytes, filename, hoja=None, fila_cabecera=None, clave="fecha"):
    """Lee a df probando filas de cabecera (la fija, o 0..4) hasta encontrar 'clave'. (df, fila)."""
    es_csv = (filename or "").lower().endswith(".csv")
    intentos = [fila_cabecera] if fila_cabecera is not None else [0, 1, 2, 3, 4]
    for h in intentos:
        try:
            if es_csv:
                df = pd.read_csv(io.BytesIO(file_bytes), header=h, dtype=str)
            else:
                df = pd.read_excel(io.BytesIO(file_bytes), sheet_name=(hoja if hoja else 0), header=h, dtype=str)
        except Exception:
            continue
        if any(clave in _norm_txt(c) for c in df.columns):
            df.columns = [str(c).strip() for c in df.columns]
            return df, (h if h is not None else 0)
    # fallback: cabecera en la fila 0
    if es_csv:
        df = pd.read_csv(io.BytesIO(file_bytes), dtype=str)
    else:
        df = pd.read_excel(io.BytesIO(file_bytes), sheet_name=(hoja if hoja else 0), dtype=str)
    df.columns = [str(c).strip() for c in df.columns]
    return df, 0


@app.post("/acd/importar")
async def acd_importar(
    x_api_key: str = Header(None),
    archivo: UploadFile = File(...),
):
    """
    Sube el Excel del ACD (resumen por agente/dia) -> acd_resumen_diario.
    Detecta la cabecera sola (formato nuevo fila 1 con Status 1..8, o viejo fila 3).
    Tiempos en segundos. Upsert por (login_acd, fecha). Cruza login_acd con agentes.
    """
    check_key(x_api_key)
    engine = get_engine()

    file_bytes = await archivo.read()
    # Detectar la fila de cabecera: formato nuevo -> fila 1 (indice 0); viejo -> fila 3 (indice 2).
    df = None
    for h in (0, 2):
        tmp = pd.read_excel(io.BytesIO(file_bytes), header=h, dtype=str)
        cols = [str(c).strip().lower() for c in tmp.columns]
        if "fecha" in cols and "agente" in cols:
            df = tmp
            break
    if df is None:
        raise HTTPException(400, "No encuentro la cabecera (columnas 'Fecha'/'Agente') en el Excel.")
    df.columns = [str(c).strip() for c in df.columns]

    mapa = {clave: _buscar_col(df.columns, prefijos) for clave, prefijos in _ACD_COLS.items()}
    faltan = [k for k in ("fecha", "login_acd") if mapa[k] is None]
    if faltan:
        raise HTTPException(400, f"Faltan columnas obligatorias {faltan}. Columnas vistas: {list(df.columns)}")

    filas = []
    for _, r in df.iterrows():
        fecha_s = _a_txt(r.get(mapa["fecha"]))
        login_s = _a_txt(r.get(mapa["login_acd"]))
        if not fecha_s or not login_s:
            continue
        try:
            fecha = pd.to_datetime(fecha_s, format="%Y%m%d").date()
        except Exception:
            continue
        filas.append({
            "fecha": fecha,
            "plataforma": _a_txt(r.get(mapa["plataforma"])) if mapa["plataforma"] else None,
            "login_acd": login_s,
            "logado": _a_hora(r.get(mapa["logado"])) if mapa["logado"] else None,
            "deslogado": _a_hora(r.get(mapa["deslogado"])) if mapa["deslogado"] else None,
            "seg_login": _a_int(r.get(mapa["seg_login"])) if mapa["seg_login"] else 0,
            "seg_not_ready": _a_int(r.get(mapa["seg_not_ready"])) if mapa["seg_not_ready"] else 0,
            "seg_pausa_visual": _a_int(r.get(mapa["seg_pausa_visual"])) if mapa["seg_pausa_visual"] else 0,
            "seg_break": _a_int(r.get(mapa["seg_break"])) if mapa["seg_break"] else 0,
            "seg_formacion": _a_int(r.get(mapa["seg_formacion"])) if mapa["seg_formacion"] else 0,
            "seg_servicios": _a_int(r.get(mapa["seg_servicios"])) if mapa["seg_servicios"] else 0,
            "seg_backoffice": _a_int(r.get(mapa["seg_backoffice"])) if mapa["seg_backoffice"] else 0,
            "seg_talk_in": _a_int(r.get(mapa["seg_talk_in"])) if mapa["seg_talk_in"] else 0,
            "seg_talk_out": _a_int(r.get(mapa["seg_talk_out"])) if mapa["seg_talk_out"] else 0,
            "seg_hold": _a_int(r.get(mapa["seg_hold"])) if mapa["seg_hold"] else 0,
            "llamadas_inbound": _a_int(r.get(mapa["llamadas_inbound"])) if mapa["llamadas_inbound"] else 0,
        })

    if not filas:
        raise HTTPException(400, "El Excel no tenia filas validas (revisa que la cabecera este en la fila 3).")

    cols = ["fecha", "plataforma", "login_acd", "logado", "deslogado", "seg_login",
            "seg_not_ready", "seg_pausa_visual", "seg_break", "seg_formacion", "seg_servicios",
            "seg_backoffice", "seg_talk_in", "seg_talk_out", "seg_hold", "llamadas_inbound"]
    actualiza = ", ".join(f"{c}=EXCLUDED.{c}" for c in cols if c not in ("login_acd", "fecha"))
    conflict = f" ON CONFLICT (login_acd, fecha) DO UPDATE SET {actualiza} "

    LOTE = 500
    try:
        with engine.begin() as conn:
            for k in range(0, len(filas), LOTE):
                lote = filas[k:k + LOTE]
                valores = []; params = {}
                for j, f in enumerate(lote):
                    valores.append("(" + ", ".join(f":{c}{j}" for c in cols) + ")")
                    for c in cols:
                        params[f"{c}{j}"] = f[c]
                sql = "INSERT INTO acd_resumen_diario (" + ", ".join(cols) + ") VALUES " + ", ".join(valores) + conflict
                conn.execute(text(sql), params)
            # Cruce login_acd -> agente_id (ambos a texto para evitar choques de tipo)
            conn.execute(text("""
                UPDATE acd_resumen_diario d
                SET agente_id = a.id
                FROM agentes a
                WHERE a.login_acd::text = d.login_acd
            """))
    except Exception as e:
        ejemplo = filas[0] if filas else {}
        raise HTTPException(500, f"Error al cargar ACD: {type(e).__name__}: {str(e)[:300]} | ejemplo: {ejemplo}")

    rep = pd.read_sql(text("""
        SELECT COUNT(*) AS filas, MIN(fecha) AS desde, MAX(fecha) AS hasta,
               COUNT(DISTINCT login_acd) AS logins,
               COUNT(DISTINCT login_acd) FILTER (WHERE agente_id IS NOT NULL) AS cruzados,
               COUNT(DISTINCT login_acd) FILTER (WHERE agente_id IS NULL) AS sin_cruzar
        FROM acd_resumen_diario
    """), engine)
    sin = pd.read_sql(text("SELECT DISTINCT login_acd FROM acd_resumen_diario WHERE agente_id IS NULL ORDER BY login_acd LIMIT 30"), engine)

    return {
        "ok": True,
        "filas_excel_validas": len(filas),
        "total_filas_tabla": int(rep["filas"][0]),
        "rango": {"desde": str(rep["desde"][0]), "hasta": str(rep["hasta"][0])},
        "logins_distintos": int(rep["logins"][0]),
        "logins_cruzados": int(rep["cruzados"][0]),
        "logins_sin_cruzar": int(rep["sin_cruzar"][0]),
        "ejemplos_sin_cruzar": sin["login_acd"].tolist(),
    }


@app.post("/conexiones/importar")
async def conexiones_importar(
    x_api_key: str = Header(None),
    archivo: UploadFile = File(...),
    campana: str = Form(...),
):
    """
    Carga GENÉRICA de conexiones del ACD -> acd_resumen_diario, guiada por la config
    'conexiones' de la campaña (campana_cargas): mapea columnas (canónico -> origen),
    detecta la fila de cabecera, parsea tiempos en segundos y horas, y cruza por login_acd.
    """
    check_key(x_api_key)
    engine = get_engine()
    cid = _id_campana(engine, campana)
    if cid is None:
        raise HTTPException(400, f"La campaña '{campana}' no existe.")
    cfg = _leer_config_carga(engine, cid, "conexiones")
    if cfg is None:
        raise HTTPException(400, f"La campaña '{campana}' no tiene config de carga 'conexiones'. Créala en el asistente.")
    cmap_cfg = cfg.get("columnas") or {}
    fmt_fecha = cfg.get("formato_fecha")  # ej '%Y%m%d'; None -> auto

    file_bytes = await archivo.read()
    df, _fila = _df_con_cabecera_auto(file_bytes, archivo.filename, cfg.get("hoja_datos"), cfg.get("fila_cabecera"))

    # mapa canónico -> columna real (a partir del origen configurado)
    mapa = {}
    for campo in CONEXIONES_CAMPOS:
        origen = cmap_cfg.get(campo)
        mapa[campo] = _buscar_col(df.columns, [origen]) if origen else None
    if not mapa.get("fecha") or not mapa.get("login_acd"):
        raise HTTPException(400, f"Faltan columnas obligatorias fecha/login_acd en la config. Columnas: {list(df.columns)}")

    def _fecha(v):
        s = _a_txt(v)
        if not s:
            return None
        try:
            return (pd.to_datetime(s, format=fmt_fecha) if fmt_fecha else pd.to_datetime(s)).date()
        except Exception:
            return None

    filas = []
    for _, r in df.iterrows():
        fecha = _fecha(r.get(mapa["fecha"]))
        login_s = _a_txt(r.get(mapa["login_acd"]))
        if not fecha or not login_s:
            continue
        fila = {
            "fecha": fecha, "login_acd": login_s,
            "plataforma": _a_txt(r.get(mapa["plataforma"])) if mapa.get("plataforma") else None,
            "logado": _a_hora(r.get(mapa["logado"])) if mapa.get("logado") else None,
            "deslogado": _a_hora(r.get(mapa["deslogado"])) if mapa.get("deslogado") else None,
        }
        for campo in _ACD_SEGUNDOS:
            fila[campo] = _a_int(r.get(mapa[campo])) if mapa.get(campo) else 0
        filas.append(fila)

    if not filas:
        raise HTTPException(400, "El archivo no tenía filas válidas (revisa el mapeo de fecha/login).")

    cols = ["fecha", "plataforma", "login_acd", "logado", "deslogado", "seg_login",
            "seg_not_ready", "seg_pausa_visual", "seg_break", "seg_formacion", "seg_servicios",
            "seg_backoffice", "seg_talk_in", "seg_talk_out", "seg_hold", "llamadas_inbound"]
    actualiza = ", ".join(f"{c}=EXCLUDED.{c}" for c in cols if c not in ("login_acd", "fecha"))
    conflict = f" ON CONFLICT (login_acd, fecha) DO UPDATE SET {actualiza} "
    LOTE = 500
    try:
        with engine.begin() as conn:
            for k in range(0, len(filas), LOTE):
                lote = filas[k:k + LOTE]
                valores = []
                params = {}
                for j, f in enumerate(lote):
                    valores.append("(" + ", ".join(f":{c}{j}" for c in cols) + ")")
                    for c in cols:
                        params[f"{c}{j}"] = f[c]
                sql = "INSERT INTO acd_resumen_diario (" + ", ".join(cols) + ") VALUES " + ", ".join(valores) + conflict
                conn.execute(text(sql), params)
            conn.execute(text("""
                UPDATE acd_resumen_diario d SET agente_id = a.id
                FROM agentes a WHERE a.login_acd::text = d.login_acd
            """))
    except Exception as e:
        raise HTTPException(500, f"Error al cargar conexiones: {type(e).__name__}: {str(e)[:300]}")

    # --- Resumen de carga SCOPED por campaña y por archivo (Hallazgo #3) ---
    # Antes este resumen (cruzados / sin cruzar / rango / logins) se calculaba sobre
    # TODA acd_resumen_diario, mezclando campañas (por eso salía "61" en vez de 56).
    # Ahora se calcula sobre los logins y fechas de ESTE archivo, cruzados contra los
    # agentes de ESTA campaña (campana_id=cid). Un login del archivo que solo cuadre
    # con un agente de OTRA campaña cuenta como "sin cruzar" aquí (lo correcto para
    # detectar errores de carga). El cruce real de agente_id en la tabla sigue siendo
    # global; su aislamiento a nivel de tabla queda para una fase posterior (Tier 2).
    logins_archivo = sorted({str(f["login_acd"]) for f in filas if f.get("login_acd")})
    fechas_archivo = [f["fecha"] for f in filas if f.get("fecha")]
    _ag = pd.read_sql(
        text("SELECT DISTINCT login_acd FROM agentes WHERE campana_id = :cid AND login_acd IS NOT NULL"),
        engine, params={"cid": cid})
    logins_campana = {str(x) for x in _ag["login_acd"].tolist()}
    cruzados = [l for l in logins_archivo if l in logins_campana]
    sin_cruzar = [l for l in logins_archivo if l not in logins_campana]
    filas_de_campana = sum(1 for f in filas if str(f.get("login_acd")) in logins_campana)
    rango_desde = str(min(fechas_archivo)) if fechas_archivo else None
    rango_hasta = str(max(fechas_archivo)) if fechas_archivo else None

    # --- A) Recalcular adherencia automáticamente para el rango recién cargado ---
    # El dashboard de adherencia lee de la tabla 'adherencia', que se llena con
    # /adherencia/calcular. Antes había que dispararlo a mano (o subir un Excel por
    # el cargador legacy /acd/importar). Aquí lo encadenamos al cargar conexiones,
    # acotado al rango del archivo. Va en try/except: si el recálculo falla, la
    # carga de conexiones NO se rompe (el usuario siempre podrá recalcular a mano).
    adherencia_recalculada = None
    try:
        _fechas = [f["fecha"] for f in filas if f.get("fecha")]
        if _fechas:
            adherencia_recalculada = adherencia_calcular(
                x_api_key=x_api_key, desde=str(min(_fechas)), hasta=str(max(_fechas)),
                campana=campana)
    except Exception as e:
        adherencia_recalculada = {"ok": False, "error": f"{type(e).__name__}: {str(e)[:200]}"}

    return {
        "ok": True, "campana": campana,
        "filas_validas": len(filas),
        "total_filas_tabla": filas_de_campana,
        "rango": {"desde": rango_desde, "hasta": rango_hasta},
        "logins_distintos": len(logins_archivo),
        "logins_cruzados": len(cruzados),
        "logins_sin_cruzar": len(sin_cruzar),
        "ejemplos_sin_cruzar": sin_cruzar[:30],
        "adherencia_recalculada": adherencia_recalculada,
    }


# ============================================================
#  ADHERENCIA · Fase 3 — Calcular adherencia (plan vs ACD)
# ============================================================
@app.post("/adherencia/calcular")
def adherencia_calcular(
    x_api_key: str = Header(None),
    desde: str = None,   # 'YYYY-MM-DD' opcional; si vacio usa el minimo del ACD
    hasta: str = None,   # 'YYYY-MM-DD' opcional; si vacio usa el maximo del ACD
    campana: str = None, # opcional: si viene, aisla rango/calculo/resumen a esa campaña
):
    """
    Cruza el plan (asignaciones, tipo='trabajo') con la realidad (acd_resumen_diario)
    por (agente_id, fecha) y calcula: ADH BRUTA, ADH NETA, UTILIZACION, PRODUCTIVIDAD,
    INFOE, mas TMO y llamadas. HR PRESENCIA sale del turno. Upsert en adherencia.

    Si se pasa 'campana', TODO el cálculo se aísla a esa campaña: el rango por defecto,
    los días que se calculan/escriben y el resumen devuelto. Sin 'campana' se comporta
    como antes (global), por compatibilidad con las llamadas existentes.
    """
    check_key(x_api_key)
    engine = get_engine()

    # Resolver campaña (opcional). Con campaña -> aislamiento por campana_id.
    cid = _id_campana(engine, campana) if campana else None
    if campana and cid is None:
        raise HTTPException(400, f"La campaña '{campana}' no existe.")

    # Rango por defecto: global si no hay campaña; acotado a la campaña si la hay
    # (vía agente_id -> agentes.campana_id, sin tocar el esquema de acd_resumen_diario).
    if cid is None:
        rango = pd.read_sql(text("SELECT MIN(fecha) AS d, MAX(fecha) AS h FROM acd_resumen_diario"), engine)
    else:
        rango = pd.read_sql(text("""
            SELECT MIN(d.fecha) AS d, MAX(d.fecha) AS h
            FROM acd_resumen_diario d
            JOIN agentes ag ON ag.id = d.agente_id
            WHERE ag.campana_id = :cid
        """), engine, params={"cid": cid})
    if rango["d"][0] is None:
        raise HTTPException(400, "No hay datos en acd_resumen_diario para ese ámbito. Importa el Excel primero.")
    d_ini = desde or str(rango["d"][0])
    d_fin = hasta or str(rango["h"][0])

    # Filtro de campaña reutilizable (asignaciones tiene campana_id).
    cond_camp = " AND a.campana_id = :cid" if cid is not None else ""
    params = {"i": d_ini, "f": d_fin}
    if cid is not None:
        params["cid"] = cid

    df = pd.read_sql(text(f"""
        SELECT a.agente_id, a.fecha, a.hora_inicio, a.hora_fin,
               d.plataforma, d.seg_login, d.seg_not_ready,
               d.seg_talk_in, d.seg_talk_out, d.seg_hold,
               d.seg_backoffice, d.llamadas_inbound
        FROM asignaciones a
        JOIN acd_resumen_diario d
          ON d.agente_id = a.agente_id AND d.fecha = a.fecha
        WHERE a.tipo = 'trabajo'
          AND a.hora_inicio IS NOT NULL AND a.hora_fin IS NOT NULL
          AND a.fecha BETWEEN :i AND :f{cond_camp}
    """), engine, params=params)

    gap = pd.read_sql(text(f"""
        SELECT COUNT(*) AS n, COUNT(DISTINCT a.agente_id) AS agentes
        FROM asignaciones a
        LEFT JOIN acd_resumen_diario d
          ON d.agente_id = a.agente_id AND d.fecha = a.fecha
        WHERE a.tipo = 'trabajo'
          AND a.hora_inicio IS NOT NULL AND a.hora_fin IS NOT NULL
          AND a.fecha BETWEEN :i AND :f{cond_camp}
          AND d.id IS NULL
    """), engine, params=params)

    if df.empty:
        return {"ok": True, "vacio": True, "rango": {"desde": d_ini, "hasta": d_fin},
                "nota": "No hay dias con plan 'trabajo' y dato ACD a la vez en ese rango.",
                "plan_sin_acd_filas": int(gap["n"][0]), "plan_sin_acd_agentes": int(gap["agentes"][0])}

    def pct(num, den):
        if not den or den <= 0:
            return None
        v = round(100.0 * num / den, 2)
        return min(v, 999.99)   # acota ruido por denominador pequeno y evita overflow numeric(6,2)

    filas = []
    for _, r in df.iterrows():
        seg_presencia = int(round(dur_turno(r["hora_inicio"], r["hora_fin"]) * 3600))
        seg_login = int(r["seg_login"] or 0)
        seg_nr = int(r["seg_not_ready"] or 0)
        seg_talk = int(r["seg_talk_in"] or 0) + int(r["seg_talk_out"] or 0)
        seg_hold = int(r["seg_hold"] or 0)
        seg_bo = int(r["seg_backoffice"] or 0)            # BO = unico estado productivo
        seg_efectiva = max(seg_login - seg_nr, 0)
        seg_productiva = seg_talk + seg_bo
        llam = int(r["llamadas_inbound"] or 0)
        filas.append({
            "agente_id": int(r["agente_id"]),
            "fecha": pd.Timestamp(r["fecha"]).date(),
            "plataforma": (str(r["plataforma"]) if r["plataforma"] is not None else None),
            "seg_presencia": seg_presencia,
            "seg_login": seg_login,
            "seg_efectiva": seg_efectiva,
            "seg_productiva": seg_productiva,
            "seg_talk": seg_talk,
            "seg_hold": seg_hold,
            "seg_not_ready": seg_nr,
            "adh_bruta": pct(seg_login, seg_presencia),         # HR Logado / HR Presencia
            "adh_neta": pct(seg_efectiva, seg_presencia),       # HR Efectiva / HR Presencia
            "utilizacion": pct(seg_efectiva, seg_login),        # HR Efectiva / HR Logado
            "productividad": pct(seg_productiva, seg_efectiva), # HR Productiva / HR Efectiva
            "infoe": pct(seg_productiva, seg_login),            # HR Productiva / HR Logado
            "llamadas": llam,
            "tmo_seg": (round((seg_talk + seg_hold) / llam, 2) if llam > 0 else None),
        })

    cols = ["agente_id", "fecha", "plataforma", "seg_presencia", "seg_login", "seg_efectiva",
            "seg_productiva", "seg_talk", "seg_hold", "seg_not_ready", "adh_bruta", "adh_neta",
            "utilizacion", "productividad", "infoe", "llamadas", "tmo_seg"]
    actualiza = ", ".join(f"{c}=EXCLUDED.{c}" for c in cols if c not in ("agente_id", "fecha"))
    conflict = f" ON CONFLICT (agente_id, fecha) DO UPDATE SET {actualiza} "

    LOTE = 500
    try:
        with engine.begin() as conn:
            for k in range(0, len(filas), LOTE):
                lote = filas[k:k + LOTE]
                valores = []; params = {}
                for j, f in enumerate(lote):
                    valores.append("(" + ", ".join(f":{c}{j}" for c in cols) + ")")
                    for c in cols:
                        params[f"{c}{j}"] = f[c]
                sql = "INSERT INTO adherencia (" + ", ".join(cols) + ") VALUES " + ", ".join(valores) + conflict
                conn.execute(text(sql), params)
    except Exception as e:
        ejemplo = filas[0] if filas else {}
        raise HTTPException(500, f"Error al escribir adherencia: {type(e).__name__}: {str(e)[:300]} | ejemplo: {ejemplo}")

    # Resumen acotado al mismo ámbito: global, o aislado a la campaña vía agentes.
    join_camp = "JOIN agentes ag ON ag.id = adh.agente_id" if cid is not None else ""
    cond_camp2 = " AND ag.campana_id = :cid" if cid is not None else ""
    rparams = {"i": d_ini, "f": d_fin}
    if cid is not None:
        rparams["cid"] = cid
    resumen = pd.read_sql(text(f"""
        SELECT COUNT(*) AS filas, COUNT(DISTINCT adh.agente_id) AS agentes,
               ROUND(100.0*SUM(adh.seg_login)   /NULLIF(SUM(adh.seg_presencia),0),1) AS adh_bruta,
               ROUND(100.0*SUM(adh.seg_efectiva)/NULLIF(SUM(adh.seg_presencia),0),1) AS adh_neta,
               ROUND(100.0*SUM(adh.seg_efectiva)/NULLIF(SUM(adh.seg_login),0),1)     AS utilizacion,
               ROUND(100.0*SUM(adh.seg_productiva)/NULLIF(SUM(adh.seg_efectiva),0),1) AS productividad,
               ROUND(100.0*SUM(adh.seg_productiva)/NULLIF(SUM(adh.seg_login),0),1)   AS infoe,
               ROUND(SUM(adh.seg_talk+adh.seg_hold)::numeric/NULLIF(SUM(adh.llamadas),0),1) AS tmo,
               SUM(adh.llamadas) AS llamadas
        FROM adherencia adh
        {join_camp}
        WHERE adh.fecha BETWEEN :i AND :f{cond_camp2}
    """), engine, params=rparams)
    rr = resumen.iloc[0]
    def fv(x):
        return None if pd.isna(x) else float(x)

    return {
        "ok": True,
        "rango": {"desde": d_ini, "hasta": d_fin},
        "filas_calculadas": len(filas),
        "agentes_cubiertos": int(rr["agentes"]),
        "adh_bruta_pct": fv(rr["adh_bruta"]),
        "adh_neta_pct": fv(rr["adh_neta"]),
        "utilizacion_pct": fv(rr["utilizacion"]),
        "productividad_pct": fv(rr["productividad"]),
        "infoe_pct": fv(rr["infoe"]),
        "tmo_seg": fv(rr["tmo"]),
        "llamadas_total": int(rr["llamadas"] or 0),
        "diagnostico_plan_sin_acd": {"filas": int(gap["n"][0]), "agentes": int(gap["agentes"][0])},
    }


# ============================================================
#  ADHERENCIA · Fase 3b — Importar parrilla historica -> asignaciones
# ============================================================
import re as _re_par

_PAR_LIBRE = {"LIBRE", "DLF", "FEST"}
_PAR_VAC = {"VAC", "VACACIONES"}
_PAR_AUS_PREFIJOS = ("BMED", "BMLD", "BAJA", "APNJ", "ANJ", "SUSP", "PER", "PERM",
                     "PR", "SNT", "SINT")


def _par_segmento(seg):
    """'09:00-18:00' -> (inicio_min, duracion_min). Tolera '.' como ':'. None si no parsea."""
    seg = seg.strip().replace(".", ":")
    m = _re_par.match(r"^(\d{1,2}):(\d{2})-(\d{1,2}):(\d{2})$", seg)
    if not m:
        return None
    h1, m1, h2, m2 = map(int, m.groups())
    if h1 > 23 or h2 > 23 or m1 > 59 or m2 > 59:
        return None
    ini = h1 * 60 + m1
    fin = h2 * 60 + m2
    dur = fin - ini
    if dur <= 0:
        dur += 24 * 60
    return ini, dur


def _par_celda(valor):
    """
    Mapea una celda de la parrilla a (tipo, hora_inicio, hora_fin).
    - turno (incl. partido con '/') -> ('trabajo', 'HH:MM', 'HH:MM')  [ventana = duracion total]
    - LIBRE/FEST/DLF -> ('libre', None, None)
    - VAC -> ('vacaciones', None, None)
    - BMED/BAJA/APNJ/SUSP/PERM... -> ('ausencia', None, None)
    - vacio -> None (no se inserta) ; no parseable -> 'SKIP'
    """
    if valor is None:
        return None
    s = str(valor).strip()
    if not s:
        return None
    up = s.upper()
    if up in _PAR_LIBRE:
        return ("libre", None, None)
    if up in _PAR_VAC:
        return ("vacaciones", None, None)
    if any(up.startswith(p) for p in _PAR_AUS_PREFIJOS):
        return ("ausencia", None, None)
    # turno (uno o varios segmentos separados por '/')
    primer_ini = None
    total_dur = 0
    for seg in s.split("/"):
        r = _par_segmento(seg)
        if r is None:
            return "SKIP"
        ini, dur = r
        if primer_ini is None:
            primer_ini = ini
        total_dur += dur
    if primer_ini is None or total_dur <= 0:
        return "SKIP"
    hi = f"{(primer_ini // 60) % 24:02d}:{primer_ini % 60:02d}"
    fin_min = (primer_ini + total_dur) % (24 * 60)
    hf = f"{fin_min // 60:02d}:{fin_min % 60:02d}"
    return ("trabajo", hi, hf)


def _mes_de_hoja(nombre):
    n = nombre.strip().lower()
    if n.startswith("abr"):
        return 4
    if n.startswith("may"):
        return 5
    if n.startswith("jun") or n.startswith("jum"):
        return 6
    return None


def _par_celda_cfg(valor, codigos):
    """
    Versión genérica de _par_celda guiada por la config 'codigos' de la campaña.
    - rango(s) horario(s) 'HH:MM-HH:MM' (también partidos con '/') -> ('trabajo', hi, hf)  [universal]
    - código de texto presente en 'codigos' -> usa su tipo (o {tipo,hora_inicio,hora_fin})
    - vacío -> None ; no reconocible -> 'SKIP'
    'codigos' es {CODE_MAYUS: "libre"|"vacaciones"|"ausencia"|"trabajo"|"ignorar"
                           | {"tipo":..., "hora_inicio":"HH:MM", "hora_fin":"HH:MM"}}
    """
    if valor is None:
        return None
    s = str(valor).strip()
    if not s:
        return None
    up = s.upper()
    if up in codigos:
        spec = codigos[up]
        if isinstance(spec, dict):
            return (spec.get("tipo", "trabajo"), spec.get("hora_inicio"), spec.get("hora_fin"))
        if str(spec).lower() == "ignorar":
            return None
        return (str(spec), None, None)
    # rango(s) horario(s) -> trabajo
    primer_ini, total_dur = None, 0
    for seg in s.split("/"):
        r = _par_segmento(seg)
        if r is None:
            return "SKIP"
        ini, dur = r
        if primer_ini is None:
            primer_ini = ini
        total_dur += dur
    if primer_ini is None or total_dur <= 0:
        return "SKIP"
    hi = f"{(primer_ini // 60) % 24:02d}:{primer_ini % 60:02d}"
    fin_min = (primer_ini + total_dur) % (24 * 60)
    hf = f"{fin_min // 60:02d}:{fin_min % 60:02d}"
    return ("trabajo", hi, hf)


def _detecta_layout_parrilla(rows, col_dni_cfg=None, fila_cab_cfg=None):
    """Detecta (fila_cabecera, col_dni, [(ci, date)...]) en una parrilla matricial."""
    if fila_cab_cfg is not None:
        fila_cab = int(fila_cab_cfg)
    else:
        fila_cab, mejor = 0, -1
        for i, r in enumerate(rows[:15]):
            nd = sum(1 for c in r if isinstance(c, datetime))
            if nd > mejor:
                mejor, fila_cab = nd, i
    header = rows[fila_cab] if fila_cab < len(rows) else ()
    cols_fecha = [(ci, c.date()) for ci, c in enumerate(header) if isinstance(c, datetime)]
    if col_dni_cfg is not None:
        col_dni = int(col_dni_cfg)
    else:
        col_dni = None
        for ci, c in enumerate(header):
            if c is not None and _norm_txt(c) in ("dni", "documento", "cedula", "nif", "id"):
                col_dni = ci
                break
        if col_dni is None:
            col_dni = 2
    return fila_cab, col_dni, cols_fecha


def _codigos_de_parrilla(rows, fila_cab, col_dni, cols_fecha, maxn=80):
    """Códigos de texto distintos (no horarios, no vacíos) hallados en las celdas de fecha."""
    cods = {}
    for r in rows[fila_cab + 1:]:
        for ci, _ in cols_fecha:
            if ci >= len(r):
                continue
            v = r[ci]
            if v is None:
                continue
            s = str(v).strip()
            if not s:
                continue
            if all(_par_segmento(seg) is not None for seg in s.split("/")):
                continue  # es un horario -> trabajo, no es "código"
            up = s.upper()
            cods[up] = cods.get(up, 0) + 1
    return [k for k, _ in sorted(cods.items(), key=lambda x: x[1], reverse=True)][:maxn]


@app.post("/asignaciones/importar-parrilla")
async def importar_parrilla(
    x_api_key: str = Header(None),
    archivo: UploadFile = File(...),
    meses: str = "4,5",            # meses a cargar (numero), separados por coma. Junio NO por defecto.
    campana: str = "Endesa",
):
    """
    Carga una parrilla matricial (1 fila por agente, 1 columna por dia) en asignaciones.
    Cruza por DNI con la tabla agentes. Por defecto solo abril y mayo (junio ya esta cargado).
    """
    check_key(x_api_key)
    engine = get_engine()

    meses_ok = {int(x) for x in str(meses).split(",") if x.strip().isdigit()}
    if not meses_ok:
        raise HTTPException(400, "Parametro 'meses' invalido (ej: '4,5').")

    id_campana = pd.read_sql(text("SELECT id FROM campanas WHERE nombre=:n"),
                             engine, params={"n": campana})
    if id_campana.empty:
        raise HTTPException(400, f"No existe la campana {campana!r}.")
    id_campana = int(id_campana["id"][0])

    # mapa DNI -> agente_id
    ag = pd.read_sql(text("SELECT id, dni FROM agentes"), engine)
    dni2id = {str(d).strip(): int(i) for i, d in zip(ag["id"], ag["dni"]) if d is not None}

    file_bytes = await archivo.read()
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)

    ID_COLS = 6          # ESTADO, Transporte, DNI, APELLIDO1, APELLIDO2, NOMBRE
    COL_DNI = 2

    filas = []
    dnis_sin_cruzar = set()
    celdas_skip = 0
    por_tipo = {"trabajo": 0, "libre": 0, "vacaciones": 0, "ausencia": 0}

    for sn in wb.sheetnames:
        mes = _mes_de_hoja(sn)
        if mes is None or mes not in meses_ok:
            continue
        ws = wb[sn]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        header = rows[0]
        # columnas de fecha de ESTE mes (cabecera tipo datetime y month == mes)
        cols_fecha = []
        for ci in range(ID_COLS, len(header)):
            h = header[ci]
            if isinstance(h, datetime) and h.month == mes:
                cols_fecha.append((ci, h.date()))
        for r in rows[1:]:
            if len(r) <= COL_DNI or r[COL_DNI] is None:
                continue
            dni = str(r[COL_DNI]).strip()
            aid = dni2id.get(dni)
            if aid is None:
                dnis_sin_cruzar.add(dni)
                continue
            for ci, fecha in cols_fecha:
                if ci >= len(r):
                    continue
                res = _par_celda(r[ci])
                if res is None:
                    continue
                if res == "SKIP":
                    celdas_skip += 1
                    continue
                tipo, hi, hf = res
                por_tipo[tipo] = por_tipo.get(tipo, 0) + 1
                filas.append({
                    "agente_id": aid, "fecha": fecha, "campana_id": id_campana,
                    "turno_id": None, "hora_inicio": hi, "hora_fin": hf,
                    "tipo": tipo, "creado_por": "import-parrilla", "bloqueado": True,
                })

    if not filas:
        raise HTTPException(400, f"No se generaron filas. Meses pedidos: {sorted(meses_ok)}. "
                                 f"DNIs sin cruzar: {len(dnis_sin_cruzar)}.")

    cols = ["agente_id", "fecha", "campana_id", "turno_id", "hora_inicio", "hora_fin",
            "tipo", "creado_por", "bloqueado"]
    actualiza = ", ".join(f"{c}=EXCLUDED.{c}" for c in cols if c not in ("agente_id", "fecha"))
    conflict = f" ON CONFLICT (agente_id, fecha) DO UPDATE SET {actualiza} "

    LOTE = 500
    try:
        with engine.begin() as conn:
            for k in range(0, len(filas), LOTE):
                lote = filas[k:k + LOTE]
                valores = []; params = {}
                for j, f in enumerate(lote):
                    valores.append("(" + ", ".join(f":{c}{j}" for c in cols) + ")")
                    for c in cols:
                        params[f"{c}{j}"] = f[c]
                sql = "INSERT INTO asignaciones (" + ", ".join(cols) + ") VALUES " + ", ".join(valores) + conflict
                conn.execute(text(sql), params)
    except Exception as e:
        ejemplo = filas[0] if filas else {}
        raise HTTPException(500, f"Error al escribir asignaciones: {type(e).__name__}: {str(e)[:300]} | ejemplo: {ejemplo}")

    return {
        "ok": True,
        "meses_cargados": sorted(meses_ok),
        "filas_insertadas": len(filas),
        "por_tipo": por_tipo,
        "agentes_cruzados": len({f["agente_id"] for f in filas}),
        "dnis_sin_cruzar": len(dnis_sin_cruzar),
        "ejemplos_dnis_sin_cruzar": sorted(dnis_sin_cruzar)[:20],
        "celdas_no_parseables": celdas_skip,
    }


@app.post("/asignaciones/importar")
async def asignaciones_importar(
    x_api_key: str = Header(None),
    archivo: UploadFile = File(...),
    campana: str = Form(...),
    meses: str = Form(None),        # opcional "4,5"; por defecto TODAS las fechas del archivo
    reemplazar: bool = Form(False),
):
    """
    Carga GENÉRICA de parrilla (turnos) -> asignaciones, guiada por la config 'turnos'
    de la campaña (campana_cargas). Detecta el layout (fila cabecera, columna DNI,
    columnas de fecha) y mapea cada celda a (tipo, hora_inicio, hora_fin) con:
      - rangos horarios 'HH:MM-HH:MM' (universal) -> trabajo
      - el diccionario 'codigos' de la config (CODE -> tipo o {tipo,hora_inicio,hora_fin})
    Cruza por DNI con la tabla agentes. Sin layout cableado por campaña.
    """
    check_key(x_api_key)
    engine = get_engine()
    cid = _id_campana(engine, campana)
    if cid is None:
        raise HTTPException(400, f"La campaña '{campana}' no existe.")
    cfg = _leer_config_carga(engine, cid, "turnos")
    if cfg is None:
        raise HTTPException(400, f"La campaña '{campana}' no tiene config de carga 'turnos'. Créala en el asistente.")
    codigos = {str(k).strip().upper(): v for k, v in (cfg.get("codigos") or {}).items()}
    col_dni_cfg = cfg.get("col_dni")
    fila_cab_cfg = cfg.get("fila_cabecera")
    meses_ok = {int(x) for x in str(meses).split(",") if x and x.strip().isdigit()} if meses else None

    ag = pd.read_sql(text("SELECT id, dni FROM agentes"), engine)
    dni2id = {str(d).strip(): int(i) for i, d in zip(ag["id"], ag["dni"]) if d is not None}

    file_bytes = await archivo.read()
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)

    filas = []
    dnis_sin_cruzar = set()
    celdas_skip = 0
    por_tipo = {}
    for sn in wb.sheetnames:
        rows = list(wb[sn].iter_rows(values_only=True))
        if not rows:
            continue
        fila_cab, col_dni, cols_fecha = _detecta_layout_parrilla(rows, col_dni_cfg, fila_cab_cfg)
        if not cols_fecha:
            continue
        if meses_ok:
            cols_fecha = [(ci, f) for ci, f in cols_fecha if f.month in meses_ok]
        for r in rows[fila_cab + 1:]:
            if len(r) <= col_dni or r[col_dni] is None:
                continue
            dni = str(r[col_dni]).strip()
            aid = dni2id.get(dni)
            if aid is None:
                dnis_sin_cruzar.add(dni)
                continue
            for ci, fecha in cols_fecha:
                if ci >= len(r):
                    continue
                res = _par_celda_cfg(r[ci], codigos)
                if res is None:
                    continue
                if res == "SKIP":
                    celdas_skip += 1
                    continue
                tipo, hi, hf = res
                por_tipo[tipo] = por_tipo.get(tipo, 0) + 1
                filas.append({
                    "agente_id": aid, "fecha": fecha, "campana_id": cid,
                    "turno_id": None, "hora_inicio": hi, "hora_fin": hf,
                    "tipo": tipo, "creado_por": "import-parrilla", "bloqueado": True,
                })

    if not filas:
        raise HTTPException(400, f"No se generaron filas. DNIs sin cruzar: {len(dnis_sin_cruzar)}. "
                                 f"Revisa que la parrilla tenga cabeceras de fecha y la columna DNI correcta.")

    cols = ["agente_id", "fecha", "campana_id", "turno_id", "hora_inicio", "hora_fin",
            "tipo", "creado_por", "bloqueado"]
    actualiza = ", ".join(f"{c}=EXCLUDED.{c}" for c in cols if c not in ("agente_id", "fecha"))
    conflict = f" ON CONFLICT (agente_id, fecha) DO UPDATE SET {actualiza} "
    LOTE = 500
    try:
        with engine.begin() as conn:
            for k in range(0, len(filas), LOTE):
                lote = filas[k:k + LOTE]
                valores = []
                params = {}
                for j, f in enumerate(lote):
                    valores.append("(" + ", ".join(f":{c}{j}" for c in cols) + ")")
                    for c in cols:
                        params[f"{c}{j}"] = f[c]
                sql = "INSERT INTO asignaciones (" + ", ".join(cols) + ") VALUES " + ", ".join(valores) + conflict
                conn.execute(text(sql), params)
    except Exception as e:
        raise HTTPException(500, f"Error al escribir asignaciones: {type(e).__name__}: {str(e)[:300]}")

    return {
        "ok": True,
        "campana": campana,
        "filas_insertadas": len(filas),
        "por_tipo": por_tipo,
        "agentes_cruzados": len({f["agente_id"] for f in filas}),
        "dnis_sin_cruzar": len(dnis_sin_cruzar),
        "ejemplos_dnis_sin_cruzar": sorted(dnis_sin_cruzar)[:20],
        "celdas_no_parseables": celdas_skip,
    }


# ============================================================
#  ADHERENCIA · Fase 4 — Datos para el dashboard
# ============================================================
@app.get("/adherencia/opciones")
def adherencia_opciones(x_api_key: str = Header(None), campana: str = None):
    """Valores para los filtros del dashboard (rango de fechas, paises, centros).
    Con 'campana' se acotan a esa campaña; sin ella, global (como antes)."""
    check_key(x_api_key)
    engine = get_engine()
    cid = _id_campana(engine, campana) if campana else None
    if campana and cid is None:
        raise HTTPException(400, f"La campaña '{campana}' no existe.")
    if cid is None:
        r = pd.read_sql(text("SELECT MIN(fecha) AS d, MAX(fecha) AS h FROM adherencia"), engine)
    else:
        r = pd.read_sql(text("""
            SELECT MIN(adh.fecha) AS d, MAX(adh.fecha) AS h
            FROM adherencia adh JOIN agentes ag ON ag.id = adh.agente_id
            WHERE ag.campana_id = :cid
        """), engine, params={"cid": cid})
    if r["d"][0] is None:
        return {"ok": True, "vacio": True}
    cond_camp = " WHERE ag.campana_id = :cid" if cid is not None else ""
    cparams = {"cid": cid} if cid is not None else {}
    pc = pd.read_sql(text(f"""
        SELECT DISTINCT ag.pais, ag.centro
        FROM adherencia adh JOIN agentes ag ON ag.id = adh.agente_id{cond_camp}
    """), engine, params=cparams)
    return {
        "ok": True, "vacio": False,
        "desde": str(r["d"][0]), "hasta": str(r["h"][0]),
        "paises": sorted([str(p) for p in pc["pais"].dropna().unique()]),
        "centros": sorted([str(c) for c in pc["centro"].dropna().unique()]),
    }


@app.get("/adherencia/dashboard")
def adherencia_dashboard(
    x_api_key: str = Header(None),
    desde: str = None,
    hasta: str = None,
    pais: str = None,
    centro: str = None,
    campana: str = None,
):
    """
    Datos agregados de adherencia para el dashboard, segun filtros.
    Adherencia y TMO se calculan ponderados por segundos/llamadas (no promedio simple).
    Con 'campana' se acota a esa campaña; sin ella, global (como antes).
    """
    check_key(x_api_key)
    engine = get_engine()

    cid = _id_campana(engine, campana) if campana else None
    if campana and cid is None:
        raise HTTPException(400, f"La campaña '{campana}' no existe.")

    if cid is None:
        rango = pd.read_sql(text("SELECT MIN(fecha) AS d, MAX(fecha) AS h FROM adherencia"), engine)
    else:
        rango = pd.read_sql(text("""
            SELECT MIN(adh.fecha) AS d, MAX(adh.fecha) AS h
            FROM adherencia adh JOIN agentes ag ON ag.id = adh.agente_id
            WHERE ag.campana_id = :cid
        """), engine, params={"cid": cid})
    if rango["d"][0] is None:
        return {"ok": True, "vacio": True}
    d_ini = desde or str(rango["d"][0])
    d_fin = hasta or str(rango["h"][0])

    cond = "adh.fecha BETWEEN :i AND :f"
    params = {"i": d_ini, "f": d_fin}
    if cid is not None:
        cond += " AND ag.campana_id = :cid"; params["cid"] = cid
    if pais:
        cond += " AND ag.pais = :pais"; params["pais"] = pais
    if centro:
        cond += " AND ag.centro = :centro"; params["centro"] = centro

    df = pd.read_sql(text(f"""
        SELECT adh.agente_id, adh.fecha, adh.seg_programados, adh.seg_adherido,
               adh.seg_talk, adh.llamadas, ag.nombre, ag.centro, ag.pais
        FROM adherencia adh
        JOIN agentes ag ON ag.id = adh.agente_id
        WHERE {cond}
    """), engine, params=params)

    if df.empty:
        return {"ok": True, "vacio": False, "sin_datos_filtro": True,
                "rango": {"desde": d_ini, "hasta": d_fin}}

    def _adh(g):
        sp = g["seg_programados"].sum()
        return round(100.0 * g["seg_adherido"].sum() / sp, 1) if sp > 0 else None

    def _tmo(g):
        ll = g["llamadas"].sum()
        return round(g["seg_talk"].sum() / ll, 1) if ll > 0 else None

    kpis = {
        "adherencia_pct": _adh(df),
        "tmo_seg": _tmo(df),
        "llamadas": int(df["llamadas"].sum()),
        "agentes": int(df["agente_id"].nunique()),
        "dias": int(df["fecha"].nunique()),
        "filas": int(len(df)),
    }

    por_fecha = []
    for fch, g in df.groupby("fecha"):
        por_fecha.append({"fecha": str(fch), "adherencia_pct": _adh(g),
                          "llamadas": int(g["llamadas"].sum()), "tmo_seg": _tmo(g)})
    por_fecha.sort(key=lambda x: x["fecha"])

    por_agente = []
    for aid, g in df.groupby("agente_id"):
        por_agente.append({
            "agente_id": int(aid), "nombre": str(g["nombre"].iloc[0]),
            "centro": (None if pd.isna(g["centro"].iloc[0]) else str(g["centro"].iloc[0])),
            "pais": (None if pd.isna(g["pais"].iloc[0]) else str(g["pais"].iloc[0])),
            "adherencia_pct": _adh(g), "tmo_seg": _tmo(g),
            "llamadas": int(g["llamadas"].sum()), "dias": int(g["fecha"].nunique()),
        })
    por_agente.sort(key=lambda x: (x["adherencia_pct"] is None, x["adherencia_pct"]))

    por_pais = []
    for ps, g in df.groupby("pais"):
        por_pais.append({"pais": str(ps), "adherencia_pct": _adh(g), "tmo_seg": _tmo(g),
                         "llamadas": int(g["llamadas"].sum()),
                         "agentes": int(g["agente_id"].nunique())})

    return {
        "ok": True, "vacio": False,
        "rango": {"desde": d_ini, "hasta": d_fin},
        "kpis": kpis,
        "por_fecha": por_fecha,
        "por_agente": por_agente,
        "por_pais": por_pais,
    }


# ============================================================
#  CONSOLA · Fase 3 — Agregar eventos de la consola -> acd_resumen_diario
# ============================================================
def _hora_local(ts, pais):
    """timestamptz -> 'HH:MM:SS' en la zona del agente. None si no hay valor."""
    if ts is None or pd.isna(ts):
        return None
    t = pd.Timestamp(ts)
    if t.tzinfo is None:
        t = t.tz_localize("UTC")
    tz = "Europe/Madrid" if str(pais) == "España" else "America/Bogota"
    try:
        return t.tz_convert(tz).strftime("%H:%M:%S")
    except Exception:
        return t.strftime("%H:%M:%S")


@app.post("/acd/agregar-consola")
def agregar_consola(
    x_api_key: str = Header(None),
    desde: str = None,   # 'YYYY-MM-DD' opcional; por defecto hoy
    hasta: str = None,
):
    """
    Resume los eventos de la consola por agente/dia y los escribe en acd_resumen_diario
    (origen='consola'), para que adherencia los calcule igual que el ACD del Excel.
    Eventos abiertos (fin NULL) se cuentan hasta ahora. Upsert por (login_acd, fecha).
    """
    check_key(x_api_key)
    engine = get_engine()

    hoy = pd.read_sql(text("SELECT (now() at time zone 'America/Bogota')::date AS d"), engine)["d"][0]
    d_ini = desde or str(hoy)
    d_fin = hasta or str(hoy)

    ev = pd.read_sql(text("""
        SELECT e.agente_id, e.fecha, e.estado,
               extract(epoch from (coalesce(e.fin, now()) - e.inicio)) AS seg,
               e.inicio, e.fin
        FROM eventos_acd e
        WHERE e.fecha BETWEEN :i AND :f
    """), engine, params={"i": d_ini, "f": d_fin})

    if ev.empty:
        return {"ok": True, "vacio": True, "rango": {"desde": d_ini, "hasta": d_fin},
                "nota": "No hay eventos de consola en ese rango."}

    ev["seg"] = ev["seg"].fillna(0).clip(lower=0).astype(int)

    ag = pd.read_sql(text("SELECT id, login_acd, pais FROM agentes"), engine)
    login_de = {int(i): (str(l) if l is not None else None) for i, l in zip(ag["id"], ag["login_acd"])}
    pais_de = {int(i): p for i, p in zip(ag["id"], ag["pais"])}

    NO_READY = {"break", "almuerzo", "descanso", "pausa_visual", "formacion",
                "no_disponible", "bano", "reunion", "gestion_bo"}

    filas = []
    sin_login = set()
    for (aid, fecha), g in ev.groupby(["agente_id", "fecha"]):
        aid = int(aid)
        login = login_de.get(aid)
        if not login:
            sin_login.add(aid)
            continue
        def s(estado):
            return int(g.loc[g["estado"] == estado, "seg"].sum())
        filas.append({
            "fecha": str(fecha),
            "plataforma": "CONSOLA",
            "login_acd": login,
            "agente_id": aid,
            "logado": _hora_local(g["inicio"].min(), pais_de.get(aid)),
            "deslogado": _hora_local(g["fin"].max(), pais_de.get(aid)),
            "seg_login": int(g["seg"].sum()),
            "seg_not_ready": int(g.loc[g["estado"].isin(NO_READY), "seg"].sum()),
            "seg_pausa_visual": s("pausa_visual"),
            "seg_break": s("break"),
            "seg_formacion": s("formacion"),
            "seg_servicios": 0,
            "seg_backoffice": s("gestion_bo"),
            "seg_talk_in": 0,     # sin telefonia aun
            "seg_talk_out": 0,
            "seg_hold": 0,
            "llamadas_inbound": 0,
        })

    if not filas:
        return {"ok": True, "vacio": False, "rango": {"desde": d_ini, "hasta": d_fin},
                "filas": 0, "agentes_sin_login_acd": sorted(sin_login),
                "nota": "Hay eventos pero ningun agente con login_acd."}

    cols = ["fecha", "plataforma", "login_acd", "agente_id", "logado", "deslogado", "seg_login",
            "seg_not_ready", "seg_pausa_visual", "seg_break", "seg_formacion", "seg_servicios",
            "seg_backoffice", "seg_talk_in", "seg_talk_out", "seg_hold", "llamadas_inbound"]
    actualiza = ", ".join(f"{c}=EXCLUDED.{c}" for c in cols if c not in ("login_acd", "fecha"))
    conflict = f" ON CONFLICT (login_acd, fecha) DO UPDATE SET {actualiza} "

    LOTE = 500
    try:
        with engine.begin() as conn:
            for k in range(0, len(filas), LOTE):
                lote = filas[k:k + LOTE]
                valores = []; params = {}
                for j, f in enumerate(lote):
                    valores.append("(" + ", ".join(f":{c}{j}" for c in cols) + ")")
                    for c in cols:
                        params[f"{c}{j}"] = f[c]
                sql = "INSERT INTO acd_resumen_diario (" + ", ".join(cols) + ") VALUES " + ", ".join(valores) + conflict
                conn.execute(text(sql), params)
            conn.execute(text("""
                UPDATE acd_resumen_diario d SET agente_id = a.id
                FROM agentes a WHERE a.login_acd::text = d.login_acd AND d.agente_id IS NULL
            """))
    except Exception as e:
        raise HTTPException(500, f"Error al agregar consola: {type(e).__name__}: {str(e)[:300]} | ejemplo: {filas[0]}")

    return {
        "ok": True,
        "rango": {"desde": d_ini, "hasta": d_fin},
        "filas_escritas": len(filas),
        "agentes": len({f["agente_id"] for f in filas}),
        "agentes_sin_login_acd": sorted(sin_login),
        "nota": "Para el dia en curso es acumulado hasta ahora; al cerrar la jornada queda completo. Ahora corre /adherencia/calcular.",
    }


# ============================================================
#  ADHERENCIA · Export a Excel (para gerencia)
# ============================================================
@app.get("/adherencia/export")
def adherencia_export(
    x_api_key: str = Header(None),
    desde: str = None,
    hasta: str = None,
    pais: str = None,
    centro: str = None,
    modo: str = None,
    supervisor: str = None,
    campana: str = None,
):
    """Descarga un .xlsx con la adherencia (detalle + resumenes) segun filtros.
    Con 'campana' se acota a esa campaña; sin ella, global (como antes)."""
    check_key(x_api_key)
    engine = get_engine()

    cid = _id_campana(engine, campana) if campana else None
    if campana and cid is None:
        raise HTTPException(400, f"La campaña '{campana}' no existe.")

    if cid is None:
        rango = pd.read_sql(text("SELECT MIN(fecha) AS d, MAX(fecha) AS h FROM adherencia"), engine)
    else:
        rango = pd.read_sql(text("""
            SELECT MIN(adh.fecha) AS d, MAX(adh.fecha) AS h
            FROM adherencia adh JOIN agentes ag ON ag.id = adh.agente_id
            WHERE ag.campana_id = :cid
        """), engine, params={"cid": cid})
    if rango["d"][0] is None:
        raise HTTPException(400, "No hay datos de adherencia. Corre /adherencia/calcular primero.")
    d_ini = desde or str(rango["d"][0])
    d_fin = hasta or str(rango["h"][0])

    cond = "adh.fecha BETWEEN :i AND :f"
    params = {"i": d_ini, "f": d_fin}
    if cid is not None: cond += " AND ag.campana_id = :cid";    params["cid"] = cid
    if pais:       cond += " AND ag.pais = :pais";              params["pais"] = pais
    if centro:     cond += " AND ag.centro = :centro";          params["centro"] = centro
    if modo:       cond += " AND ag.modo = :modo";              params["modo"] = modo
    if supervisor: cond += " AND ag.supervisor_nombre = :sup";  params["sup"] = supervisor

    df = pd.read_sql(text(f"""
        SELECT adh.fecha, ag.nombre, ag.dni, ag.centro, ag.pais, ag.modo, ag.supervisor_nombre,
               adh.seg_presencia, adh.seg_login, adh.seg_efectiva, adh.seg_productiva,
               adh.seg_talk, adh.seg_hold, adh.llamadas, adh.tmo_seg,
               adh.adh_bruta, adh.adh_neta, adh.utilizacion, adh.productividad, adh.infoe
        FROM adherencia adh JOIN agentes ag ON ag.id = adh.agente_id
        WHERE {cond}
        ORDER BY ag.nombre, adh.fecha
    """), engine, params=params)
    if df.empty:
        raise HTTPException(404, "No hay datos para esos filtros.")

    def kpis(g):
        sp, sl, se = g["seg_presencia"].sum(), g["seg_login"].sum(), g["seg_efectiva"].sum()
        spr, ll = g["seg_productiva"].sum(), g["llamadas"].sum()
        f = lambda n, d: (round(100.0 * n / d, 1) if d > 0 else None)
        return pd.Series({
            "HR Presencia": round(sp / 3600, 1), "HR Logado": round(sl / 3600, 1),
            "HR Efectiva": round(se / 3600, 1), "HR Productiva": round(spr / 3600, 1),
            "ADH Bruta %": f(sl, sp), "ADH Neta %": f(se, sp), "Utilizacion %": f(se, sl),
            "Productividad %": f(spr, se), "INFOE %": f(spr, sl),
            "Llamadas": int(ll),
            "TMO (mm:ss)": (f"{int((g['seg_talk'].sum()+g['seg_hold'].sum())//ll)//60:02d}:"
                            f"{int((g['seg_talk'].sum()+g['seg_hold'].sum())//ll)%60:02d}" if ll > 0 else ""),
            "Dias": g["fecha"].nunique(),
        })

    # Detalle por agente/dia
    det = df.copy()
    det["HR Presencia"] = (det["seg_presencia"] / 3600).round(2)
    det["HR Logado"] = (det["seg_login"] / 3600).round(2)
    det["HR Efectiva"] = (det["seg_efectiva"] / 3600).round(2)
    det["HR Productiva"] = (det["seg_productiva"] / 3600).round(2)
    detalle = det[["fecha", "nombre", "dni", "centro", "pais", "modo", "supervisor_nombre",
                   "HR Presencia", "HR Logado", "HR Efectiva", "HR Productiva",
                   "adh_bruta", "adh_neta", "utilizacion", "productividad", "infoe",
                   "llamadas", "tmo_seg"]].rename(columns={
        "fecha": "Fecha", "nombre": "Agente", "dni": "DNI", "centro": "Centro", "pais": "Pais",
        "modo": "Modo", "supervisor_nombre": "Supervisor", "adh_bruta": "ADH Bruta %",
        "adh_neta": "ADH Neta %", "utilizacion": "Utilizacion %", "productividad": "Productividad %",
        "infoe": "INFOE %", "llamadas": "Llamadas", "tmo_seg": "TMO (seg)"})

    res_ag = df.groupby(["nombre", "centro", "pais", "modo", "supervisor_nombre"]).apply(kpis).reset_index()
    res_ag = res_ag.rename(columns={"nombre": "Agente", "centro": "Centro", "pais": "Pais",
                                    "modo": "Modo", "supervisor_nombre": "Supervisor"})
    res_pais = df.groupby("pais").apply(kpis).reset_index().rename(columns={"pais": "Pais"})
    res_modo = df.groupby("modo").apply(kpis).reset_index().rename(columns={"modo": "Modo"})
    general = kpis(df).to_frame(name=f"{d_ini} a {d_fin}").reset_index().rename(columns={"index": "Indicador"})

    import io as _io
    buf = _io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as xw:
        general.to_excel(xw, sheet_name="Resumen general", index=False)
        res_ag.to_excel(xw, sheet_name="Por agente", index=False)
        res_pais.to_excel(xw, sheet_name="Por pais", index=False)
        res_modo.to_excel(xw, sheet_name="Por modo", index=False)
        detalle.to_excel(xw, sheet_name="Detalle", index=False)
    buf.seek(0)

    from fastapi.responses import StreamingResponse
    nombre = f"adherencia_{d_ini}_a_{d_fin}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{nombre}"'},
    )


# ============================================================
#  CRM · Importar base de contactos (xlsx/csv) -> crm_contactos
# ============================================================
@app.post("/crm/importar")
async def crm_importar(
    x_api_key: str = Header(None),
    archivo: UploadFile = File(...),
    campana: str = "Endesa",
):
    """Sube una base de contactos (xlsx o csv). Columnas conocidas -> campos; el resto -> datos_extra (jsonb)."""
    check_key(x_api_key)
    engine = get_engine()
    import json as _json

    raw = await archivo.read()
    nombre_arch = (archivo.filename or "").lower()
    if nombre_arch.endswith(".csv"):
        df = pd.read_csv(io.BytesIO(raw), dtype=str)
    else:
        df = pd.read_excel(io.BytesIO(raw), dtype=str)
    df.columns = [str(c).strip() for c in df.columns]

    cc = pd.read_sql(text("SELECT id FROM campanas WHERE nombre=:n"), engine, params={"n": campana})
    id_camp = int(cc["id"][0]) if not cc.empty else None

    conocidas = {"nombre", "documento", "telefono", "telefono2", "email", "direccion", "ciudad"}
    real = {c.lower(): c for c in df.columns}
    extra_cols = [c for c in df.columns if c.lower() not in conocidas]

    def val(r, key):
        c = real.get(key)
        if c is None:
            return None
        v = r.get(c)
        if v is None or (isinstance(v, float) and pd.isna(v)):
            return None
        s = str(v).strip()
        return s if s and s.lower() != "nan" else None

    filas = []
    for _, r in df.iterrows():
        extra = {}
        for c in extra_cols:
            v = r.get(c)
            if v is not None and not (isinstance(v, float) and pd.isna(v)):
                s = str(v).strip()
                if s and s.lower() != "nan":
                    extra[c] = s
        filas.append({
            "nombre": val(r, "nombre"), "documento": val(r, "documento"),
            "telefono": val(r, "telefono"), "telefono2": val(r, "telefono2"),
            "email": val(r, "email"), "direccion": val(r, "direccion"), "ciudad": val(r, "ciudad"),
            "campana_id": id_camp, "estado": "pendiente",
            "datos_extra": (_json.dumps(extra, ensure_ascii=False) if extra else None),
        })
    if not filas:
        raise HTTPException(400, "El archivo no tenia filas.")

    cols = ["nombre", "documento", "telefono", "telefono2", "email", "direccion", "ciudad",
            "campana_id", "estado", "datos_extra"]
    LOTE = 500
    insertadas = 0
    try:
        with engine.begin() as conn:
            for k in range(0, len(filas), LOTE):
                lote = filas[k:k + LOTE]
                valores = []; params = {}
                for j, f in enumerate(lote):
                    ph = []
                    for c in cols:
                        ph.append(f"cast(:{c}{j} as jsonb)" if c == "datos_extra" else f":{c}{j}")
                        params[f"{c}{j}"] = f[c]
                    valores.append("(" + ", ".join(ph) + ")")
                sql = "INSERT INTO crm_contactos (" + ", ".join(cols) + ") VALUES " + ", ".join(valores)
                conn.execute(text(sql), params)
                insertadas += len(lote)
    except Exception as e:
        raise HTTPException(500, f"Error al importar contactos: {type(e).__name__}: {str(e)[:300]} | ejemplo: {filas[0]}")

    return {"ok": True, "contactos_importados": insertadas,
            "columnas_extra_a_datos_extra": extra_cols, "campana": campana}


# ============================================================
#  Agentes · Importar horas/salario por Excel (cruza por dni)
# ============================================================
@app.post("/agentes/entra-roster")
async def agentes_entra_roster(
    x_api_key: str = Header(None),
    ids: str = Form(...),            # IDs de agente separados por coma: "10" o "10,20,30"
    entra_roster: bool = Form(...),  # valor a aplicar (true = entra al roster)
    campana: str = Form(None),       # opcional: acota a la campaña (seguridad multi-campaña)
):
    """Marca/desmarca 'entra_roster' para uno o varios agentes (edición individual o masiva)."""
    check_key(x_api_key)
    engine = get_engine()
    try:
        id_list = [int(x.strip()) for x in str(ids).split(",") if str(x).strip()]
    except ValueError:
        raise HTTPException(400, "ids debe ser una lista de enteros separados por coma.")
    if not id_list:
        raise HTTPException(400, "No se recibieron ids de agentes.")
    in_sql, params = _in_clause("id", id_list)
    params["val"] = bool(entra_roster)
    cond = f"id IN {in_sql}"
    if campana:
        cid = _id_campana(engine, campana)
        if cid is None:
            raise HTTPException(400, f"La campaña '{campana}' no existe.")
        cond += " AND campana_id = :cid"
        params["cid"] = int(cid)
    with engine.begin() as conn:
        res = conn.execute(text(f"UPDATE agentes SET entra_roster = :val WHERE {cond}"), params)
        n = res.rowcount
    return {"ok": True, "actualizados": int(n), "entra_roster": bool(entra_roster)}


@app.post("/agentes/importar-horas")
async def agentes_importar_horas(
    x_api_key: str = Header(None),
    archivo: UploadFile = File(...),
):
    """Sube xlsx/csv con columna 'dni' y 'horas_semanales' y/o 'salario_mensual'. Actualiza agentes por dni."""
    check_key(x_api_key)
    engine = get_engine()
    raw = await archivo.read()
    nombre = (archivo.filename or "").lower()
    if nombre.endswith(".csv"):
        df = pd.read_csv(io.BytesIO(raw), dtype=str)
    else:
        df = pd.read_excel(io.BytesIO(raw), dtype=str)
    df.columns = [str(c).strip().lower() for c in df.columns]

    if "dni" not in df.columns:
        raise HTTPException(400, "El archivo debe tener una columna 'dni'.")
    tiene_h = "horas_semanales" in df.columns
    tiene_s = "salario_mensual" in df.columns
    if not tiene_h and not tiene_s:
        raise HTTPException(400, "El archivo debe tener 'horas_semanales' y/o 'salario_mensual'.")

    def num(v):
        if v is None:
            return None
        s = str(v).strip().replace(".", "").replace(",", ".") if False else str(v).strip().replace(",", ".")
        if s == "" or s.lower() == "nan":
            return None
        try:
            return float(s)
        except Exception:
            return None

    filas = []
    for _, r in df.iterrows():
        dni = str(r.get("dni")).strip() if r.get("dni") is not None else None
        if not dni or dni.lower() == "nan":
            continue
        filas.append({
            "dni": dni,
            "h": num(r.get("horas_semanales")) if tiene_h else None,
            "s": num(r.get("salario_mensual")) if tiene_s else None,
        })
    if not filas:
        raise HTTPException(400, "No se encontraron filas con dni.")

    valores = []; params = {}
    for j, f in enumerate(filas):
        valores.append(f"(:dni{j}, cast(:h{j} as numeric), cast(:s{j} as numeric))")
        params[f"dni{j}"] = f["dni"]; params[f"h{j}"] = f["h"]; params[f"s{j}"] = f["s"]
    sql = f"""
        update agentes a set
          horas_semanales = coalesce(v.h, a.horas_semanales),
          salario_mensual = coalesce(v.s, a.salario_mensual)
        from (values {", ".join(valores)}) as v(dni, h, s)
        where a.dni = v.dni
    """
    with engine.begin() as conn:
        res = conn.execute(text(sql), params)
        actualizados = res.rowcount or 0

    return {"ok": True, "filas_archivo": len(filas), "agentes_actualizados": actualizados,
            "sin_cruce": len(filas) - actualizados}


# ============================================================
#  Nómina · Calcular horas y recargos por rango -> nomina_dia
# ============================================================
@app.post("/nomina/calcular")
def nomina_calcular(
    desde: str,
    hasta: str,
    x_api_key: str = Header(None),
):
    """Calcula nomina_dia (horas + recargos) para el rango [desde, hasta]. Borra y recalcula ese rango.
    CO: franja nocturna 19-06 y $ (valor_hora = salario/240, festivo 80/90/100 según fecha).
    ES: franja nocturna 22-06, solo horas (sin $)."""
    check_key(x_api_key)
    engine = get_engine()
    from datetime import datetime, timedelta, date as _date, time as _time
    from zoneinfo import ZoneInfo
    TZ_MAD = ZoneInfo("Europe/Madrid"); TZ_BOG = ZoneInfo("America/Bogota")

    datetime.strptime(desde, "%Y-%m-%d"); datetime.strptime(hasta, "%Y-%m-%d")  # valida formato

    def norm_pais(p):
        u = (p or "").strip().upper()
        if u in ("CO", "COL", "COLOMBIA"): return "CO"
        if u in ("ES", "ESP", "ESPANA", "ESPAÑA"): return "ES"
        return u

    def _cu(c):
        if c is None:
            return None
        s = str(c).strip().upper()
        return s if s and s != "NAN" else None

    ag = pd.read_sql(text("select id, pais, salario_mensual, centro from agentes"), engine)
    info = {int(r.id): (norm_pais(r.pais),
                        (float(r.salario_mensual) if pd.notna(r.salario_mensual) else None),
                        _cu(r.centro))
            for r in ag.itertuples()}

    fe = pd.read_sql(text("select fecha, pais, centro from festivos"), engine)
    fest_pais = {}    # festivos de todo el país (centro nulo)
    fest_centro = {}  # festivos por centro: (pais, centro) -> set(fechas)
    for r in fe.itertuples():
        d = r.fecha if isinstance(r.fecha, _date) else pd.to_datetime(r.fecha).date()
        p = norm_pais(r.pais); c = _cu(r.centro)
        if c is None:
            fest_pais.setdefault(p, set()).add(d)
        else:
            fest_centro.setdefault((p, c), set()).add(d)

    asg = pd.read_sql(text("""
        select agente_id, fecha, hora_inicio, hora_fin, tipo, pago
        from asignaciones where fecha between :d and :h
    """), engine, params={"d": desde, "h": hasta})

    DIV = 240.0
    def factor_festivo(d):
        if d >= _date(2027, 7, 1): return 1.00
        if d >= _date(2026, 7, 1): return 0.90
        if d >= _date(2025, 7, 1): return 0.80
        return 0.75

    def overlap_h(s, e, ws, we):
        lo = max(s, ws); hi = min(e, we)
        return max(0.0, (hi - lo).total_seconds()) / 3600.0

    def to_time(x):
        if isinstance(x, _time): return x
        s = str(x)
        parts = s.split(":")
        return _time(int(parts[0]), int(parts[1]), int(parts[2]) if len(parts) > 2 else 0)

    filas = []
    for r in asg.itertuples():
        aid = int(r.agente_id)
        pais, salario, centro = info.get(aid, ("", None, None))
        fecha = r.fecha if isinstance(r.fecha, _date) else pd.to_datetime(r.fecha).date()
        tipo = (r.tipo or "").strip()
        pago = bool(r.pago) if r.pago is not None else True
        base = dict(agente_id=aid, fecha=fecha.isoformat(), tipo=tipo, pago=pago,
                    horas_prog=0.0, h_diurna_ord=0.0, h_nocturna=0.0,
                    h_festiva_diurna=0.0, h_festiva_nocturna=0.0,
                    val_rec_nocturno=0.0, val_rec_festivo=0.0, val_rec_noct_festivo=0.0)
        es_trabajo = (tipo == "trabajo") and (r.hora_inicio is not None) and (r.hora_fin is not None)
        if not es_trabajo:
            filas.append(base); continue

        ti = to_time(r.hora_inicio); tf = to_time(r.hora_fin)
        start = datetime.combine(fecha, ti); end = datetime.combine(fecha, tf)
        if end <= start: end += timedelta(days=1)
        base["horas_prog"] = round((end - start).total_seconds() / 3600.0, 4)

        # Las horas vienen en hora de Madrid. Para agentes CO se convierten a hora de Bogotá
        # (maneja verano/invierno automáticamente) y se clasifican en hora local colombiana.
        if pais == "CO":
            start = start.replace(tzinfo=TZ_MAD).astimezone(TZ_BOG).replace(tzinfo=None)
            end   = end.replace(tzinfo=TZ_MAD).astimezone(TZ_BOG).replace(tzinfo=None)

        day_end_hour = 19 if pais == "CO" else 22
        cur = start
        while cur < end:
            D = cur.date()
            seg_e = min(end, datetime.combine(D + timedelta(days=1), _time(0, 0)))
            ws = datetime.combine(D, _time(6, 0)); we = datetime.combine(D, _time(day_end_hour, 0))
            h_total = (seg_e - cur).total_seconds() / 3600.0
            h_dia = overlap_h(cur, seg_e, ws, we)
            h_noc = max(0.0, h_total - h_dia)
            es_fest = (D.weekday() == 6) or (D in fest_pais.get(pais, set())) \
                or (centro is not None and D in fest_centro.get((pais, centro), set()))
            if es_fest:
                base["h_festiva_diurna"] += h_dia; base["h_festiva_nocturna"] += h_noc
                if pais == "CO" and salario:
                    vh = salario / DIV; ff = factor_festivo(D)
                    base["val_rec_festivo"] += h_dia * vh * ff
                    base["val_rec_noct_festivo"] += h_noc * vh * (0.35 + ff)
            else:
                base["h_diurna_ord"] += h_dia; base["h_nocturna"] += h_noc
                if pais == "CO" and salario:
                    base["val_rec_nocturno"] += h_noc * (salario / DIV) * 0.35
            cur = seg_e

        for k in ["h_diurna_ord", "h_nocturna", "h_festiva_diurna", "h_festiva_nocturna"]:
            base[k] = round(base[k], 4)
        for k in ["val_rec_nocturno", "val_rec_festivo", "val_rec_noct_festivo"]:
            base[k] = round(base[k], 2)
        filas.append(base)

    cols = ["agente_id", "fecha", "tipo", "pago", "horas_prog", "h_diurna_ord", "h_nocturna",
            "h_festiva_diurna", "h_festiva_nocturna", "val_rec_nocturno", "val_rec_festivo", "val_rec_noct_festivo"]
    with engine.begin() as conn:
        conn.execute(text("delete from nomina_dia where fecha between :d and :h"), {"d": desde, "h": hasta})
        LOTE = 500
        for k in range(0, len(filas), LOTE):
            lote = filas[k:k + LOTE]
            vals = []; params = {}
            for j, f in enumerate(lote):
                vals.append("(" + ",".join(f":{c}{j}" for c in cols) + ")")
                for c in cols:
                    params[f"{c}{j}"] = f[c]
            conn.execute(text("insert into nomina_dia (" + ",".join(cols) + ") values " + ",".join(vals)), params)

    return {"ok": True, "rango": [desde, hasta], "filas_calculadas": len(filas),
            "agentes": len(set(f["agente_id"] for f in filas))}


# ============================================================
#  Vacaciones · Calcular saldos por año de servicio -> vacaciones_saldo
# ============================================================
@app.post("/vacaciones/calcular")
def vacaciones_calcular(x_api_key: str = Header(None)):
    """Calcula el saldo de vacaciones por agente según su año de servicio (aniversario de fecha_alta).
    CO: 15 hábiles (excluye domingos y festivos). ES: 23 laborables (excluye sáb, dom y festivos).
    Días tomados = asignaciones tipo='vacaciones' dentro del periodo (incluye programadas a futuro)."""
    check_key(x_api_key)
    engine = get_engine()
    from datetime import datetime, timedelta, date as _date
    from zoneinfo import ZoneInfo

    hoy = datetime.now(ZoneInfo("America/Bogota")).date()

    def norm_pais(p):
        u = (p or "").strip().upper()
        if u in ("CO", "COL", "COLOMBIA"): return "CO"
        if u in ("ES", "ESP", "ESPANA", "ESPAÑA"): return "ES"
        return u

    def _cu(c):
        if c is None: return None
        s = str(c).strip().upper()
        return s if s and s != "NAN" else None

    def aniversario(fa, anio):
        try:
            return fa.replace(year=anio)
        except ValueError:
            return fa.replace(year=anio, month=2, day=28)  # 29-feb -> 28-feb

    ag = pd.read_sql(text("select id, pais, centro, fecha_alta, vacaciones_anuales from agentes"), engine)

    fe = pd.read_sql(text("select fecha, pais, centro from festivos"), engine)
    fest_pais, fest_centro = {}, {}
    for r in fe.itertuples():
        d = r.fecha if isinstance(r.fecha, _date) else pd.to_datetime(r.fecha).date()
        p = norm_pais(r.pais); c = _cu(r.centro)
        (fest_pais.setdefault(p, set()) if c is None else fest_centro.setdefault((p, c), set())).add(d)

    av = pd.read_sql(text("select agente_id, fecha from asignaciones where tipo = 'vacaciones'"), engine)
    vac_por_agente = {}
    for r in av.itertuples():
        d = r.fecha if isinstance(r.fecha, _date) else pd.to_datetime(r.fecha).date()
        vac_por_agente.setdefault(int(r.agente_id), []).append(d)

    def es_festivo(d, pais, centro):
        return (d in fest_pais.get(pais, set())) or (centro is not None and d in fest_centro.get((pais, centro), set()))

    filas = []
    for r in ag.itertuples():
        aid = int(r.id)
        pais = norm_pais(r.pais); centro = _cu(r.centro)
        fa = r.fecha_alta if isinstance(r.fecha_alta, _date) else (pd.to_datetime(r.fecha_alta).date() if pd.notna(r.fecha_alta) else None)
        anuales = float(r.vacaciones_anuales) if pd.notna(r.vacaciones_anuales) else None
        if fa is None or anuales is None:
            continue

        if pais == "ES":
            # España: año calendario, cupo completo disponible desde el 1 de enero
            periodo_inicio = _date(hoy.year, 1, 1)
            periodo_fin = _date(hoy.year, 12, 31)
            causado = anuales
        else:
            # Colombia: año de servicio (aniversario de fecha_alta), causado proporcional
            aniv = aniversario(fa, hoy.year)
            periodo_inicio = aniv if aniv <= hoy else aniversario(fa, hoy.year - 1)
            periodo_fin = aniversario(periodo_inicio, periodo_inicio.year + 1) - timedelta(days=1)
            dias_periodo = (periodo_fin - periodo_inicio).days + 1
            dias_transc = (min(hoy, periodo_fin) - periodo_inicio).days + 1
            causado = round(anuales * dias_transc / dias_periodo, 1)

        tomados = 0
        for d in vac_por_agente.get(aid, []):
            if not (periodo_inicio <= d <= periodo_fin):
                continue
            if pais == "CO":
                if d.weekday() == 6 or es_festivo(d, pais, centro):  # domingo o festivo
                    continue
            else:  # ES: laborables -> excluye sábado, domingo y festivos
                if d.weekday() >= 5 or es_festivo(d, pais, centro):
                    continue
            tomados += 1

        filas.append(dict(
            agente_id=aid, periodo_inicio=periodo_inicio.isoformat(), periodo_fin=periodo_fin.isoformat(),
            unidad=("hábiles" if pais == "CO" else "laborables"),
            cupo_anual=anuales, causado=causado, tomados=tomados,
            saldo_disponible=round(causado - tomados, 1), pendiente_cupo=round(anuales - tomados, 1),
        ))

    cols = ["agente_id", "periodo_inicio", "periodo_fin", "unidad", "cupo_anual",
            "causado", "tomados", "saldo_disponible", "pendiente_cupo"]
    with engine.begin() as conn:
        conn.execute(text("delete from vacaciones_saldo"))
        LOTE = 500
        for k in range(0, len(filas), LOTE):
            lote = filas[k:k + LOTE]
            vals = []; params = {}
            for j, f in enumerate(lote):
                vals.append("(" + ",".join(f":{c}{j}" for c in cols) + ")")
                for c in cols:
                    params[f"{c}{j}"] = f[c]
            conn.execute(text("insert into vacaciones_saldo (" + ",".join(cols) + ") values " + ",".join(vals)), params)

    return {"ok": True, "agentes": len(filas), "fecha_referencia": hoy.isoformat()}


# ============================================================
#  Vacaciones · Importar plantilla (hoja APROBADAS) -> fecha_alta + asignaciones
# ============================================================
@app.post("/vacaciones/importar")
async def vacaciones_importar(
    x_api_key: str = Header(None),
    archivo: UploadFile = File(...),
    hoja: str = "APROBADAS",
):
    """Lee la plantilla de vacaciones (hoja APROBADAS): col B=DNI, H=FECHA ANTIGÜEDAD,
    N..AN=fechas aprobadas. Actualiza agentes.fecha_alta y carga las fechas como
    asignaciones tipo='vacaciones' (si ese día había 'trabajo', lo convierte)."""
    check_key(x_api_key)
    engine = get_engine()
    from openpyxl import load_workbook
    from datetime import datetime, date as _date

    raw = await archivo.read()
    try:
        wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    except Exception as e:
        raise HTTPException(400, f"No pude abrir el Excel: {type(e).__name__}. Si tiene contraseña de archivo, quítala y vuelve a subirlo.")
    if hoja not in wb.sheetnames:
        raise HTTPException(400, f"No existe la hoja '{hoja}'. Hojas: {wb.sheetnames}")
    ws = wb[hoja]

    ag = pd.read_sql(text("select id, dni, campana_id from agentes"), engine)
    por_dni = {str(r.dni).strip().upper(): (int(r.id), (int(r.campana_id) if pd.notna(r.campana_id) else None))
               for r in ag.itertuples() if r.dni is not None}

    altas, vacs, sin_cruce = [], [], []
    for row in ws.iter_rows(min_row=3, max_col=40, values_only=True):
        dni = row[1]
        if dni in (None, ""):
            continue
        dni = str(dni).strip().upper()
        cruce = por_dni.get(dni)
        if cruce is None:
            sin_cruce.append(dni)
            continue
        aid, camp = cruce
        fa = row[7]
        turno = str(row[4]).strip().upper() if row[4] not in (None, "") else None
        if isinstance(fa, datetime):
            altas.append({"aid": aid, "fa": fa.date().isoformat(), "turno": turno})
        for v in row[13:40]:
            if isinstance(v, datetime):
                vacs.append({"aid": aid, "fecha": v.date().isoformat(), "camp": camp})

    if not altas and not vacs:
        raise HTTPException(400, "No encontré datos (¿la hoja tiene el formato esperado?).")

    existentes = pd.read_sql(text("select id, agente_id, fecha, tipo from asignaciones"), engine)
    mapa = {}
    for r in existentes.itertuples():
        f = r.fecha if isinstance(r.fecha, _date) else pd.to_datetime(r.fecha).date()
        mapa[(int(r.agente_id), f.isoformat())] = (int(r.id), (r.tipo or "").strip())

    nuevos, convertir, ya_estaban = [], [], 0
    for v in vacs:
        key = (v["aid"], v["fecha"])
        ex = mapa.get(key)
        if ex is None:
            nuevos.append(v)
        elif ex[1] == "vacaciones":
            ya_estaban += 1
        else:
            convertir.append(ex[0])

    with engine.begin() as conn:
        # fecha_alta real
        if altas:
            vals = []; params = {}
            for j, a in enumerate(altas):
                vals.append(f"(:a{j}, cast(:f{j} as date), :t{j})")
                params[f"a{j}"] = a["aid"]; params[f"f{j}"] = a["fa"]; params[f"t{j}"] = a["turno"]
            conn.execute(text(
                "update agentes g set fecha_alta = v.fa, turno = coalesce(v.turno, g.turno) "
                "from (values " + ",".join(vals) +
                ") as v(aid, fa, turno) where g.id = v.aid"), params)
        # convertir trabajo/libre/ausencia de ese día en vacaciones
        if convertir:
            conn.execute(text(
                "update asignaciones set tipo='vacaciones', hora_inicio=null, hora_fin=null, pago=true "
                "where id = any(:ids)"), {"ids": convertir})
        # insertar las que no existían
        LOTE = 500
        for k in range(0, len(nuevos), LOTE):
            lote = nuevos[k:k + LOTE]
            vals = []; params = {}
            for j, n in enumerate(lote):
                vals.append(f"(:a{j}, cast(:f{j} as date), 'vacaciones', true, :c{j}, 'import_vacaciones')")
                params[f"a{j}"] = n["aid"]; params[f"f{j}"] = n["fecha"]; params[f"c{j}"] = n["camp"]
            conn.execute(text(
                "insert into asignaciones (agente_id, fecha, tipo, pago, campana_id, creado_por) values "
                + ",".join(vals)), params)

    return {"ok": True, "hoja": hoja,
            "fechas_alta_actualizadas": len(altas),
            "vacaciones_insertadas": len(nuevos),
            "convertidas_de_otro_tipo": len(convertir),
            "ya_existian": ya_estaban,
            "dnis_sin_cruce": sorted(set(sin_cruce))}


# ============================================================
#  Agentes · Alta/actualización masiva por campaña (xlsx/csv)
# ============================================================
@app.post("/agentes/importar")
async def agentes_importar(
    campana: str,
    x_api_key: str = Header(None),
    archivo: UploadFile = File(...),
):
    """Monta o actualiza los agentes de una campaña. Cruza por 'dni': si existe lo actualiza
    (solo campos con valor) y lo asigna a la campaña; si no, lo crea. Crea la campaña si no existe.
    Crea también las filas de 'usuarios' (<dni>@agentes.wfm, rol agente).
    Columnas: dni*, nombre, centro, pais, modo, turno, login_acd, fecha_alta,
    jornada_horas, salario_mensual, vacaciones_anuales."""
    check_key(x_api_key)
    engine = get_engine()
    from datetime import datetime, date as _date

    raw = await archivo.read()
    nombre_arch = (archivo.filename or "").lower()

    # ¿La campaña tiene config de 'agentes'? -> mapea las columnas del archivo a los nombres canónicos.
    _cid = pd.read_sql(text("SELECT id FROM campanas WHERE nombre=:n"), engine, params={"n": campana})
    _cfg_ag = _leer_config_carga(engine, int(_cid["id"][0]), "agentes") if not _cid.empty else None
    _hoja_ag = (_cfg_ag or {}).get("hoja_datos")

    if nombre_arch.endswith(".csv"):
        df = pd.read_csv(io.BytesIO(raw), dtype=str)
    else:
        df = pd.read_excel(io.BytesIO(raw), sheet_name=(_hoja_ag if _hoja_ag else 0), dtype=str)
    df.columns = [str(c).strip().lower() for c in df.columns]

    # Renombrar según el mapeo de la config (origen -> canónico). Si no hay config, se usan
    # los nombres tal cual vengan (compatibilidad con el formato clásico de Endesa).
    if _cfg_ag and isinstance(_cfg_ag.get("columnas"), dict):
        ren = {}
        for canon, origen in _cfg_ag["columnas"].items():
            if origen:
                ren[str(origen).strip().lower()] = canon
        df = df.rename(columns=ren)

    if "dni" not in df.columns:
        raise HTTPException(400, "El archivo debe tener una columna 'dni' (o mapéala en el asistente).")

    def S(v):
        if v is None: return None
        s = str(v).strip()
        return s if s and s.lower() != "nan" else None

    def N(v):
        s = S(v)
        if s is None: return None
        try: return float(s.replace(",", "."))
        except Exception: return None

    def F(v):
        s = S(v)
        if s is None: return None
        try: return pd.to_datetime(s).date().isoformat()
        except Exception: return None

    def pais_norm(p):
        s = (S(p) or "").upper()
        if s in ("CO", "COL", "COLOMBIA"): return "Colombia"
        if s in ("ES", "ESP", "ESPANA", "ESPAÑA"): return "España"
        return S(p)

    def roster_norm(v):
        # sí/no/1/0/x/true/false -> bool. Si no viene o no se entiende -> None
        # (None: en alta usa default true; en update conserva el valor actual).
        s = S(v)
        if s is None: return None
        s = s.strip().lower()
        if s in ("si", "sí", "true", "1", "x", "yes", "y", "verdadero", "v"): return True
        if s in ("no", "false", "0", "n", "falso", "f"): return False
        return None

    filas = []
    for _, r in df.iterrows():
        dni = S(r.get("dni"))
        if not dni: continue
        p = pais_norm(r.get("pais"))
        vac = N(r.get("vacaciones_anuales"))
        if vac is None and p == "Colombia": vac = 15
        if vac is None and p == "España":   vac = 23
        filas.append({
            "dni": dni.upper(), "nombre": S(r.get("nombre")),
            "centro": (S(r.get("centro")) or "").upper() or None,
            "pais": p, "modo": (S(r.get("modo")) or "").upper() or None,
            "turno": (S(r.get("turno")) or "").upper() or None,
            "login_acd": S(r.get("login_acd")), "fecha_alta": F(r.get("fecha_alta")),
            "jornada_horas": N(r.get("jornada_horas")),
            "salario_mensual": N(r.get("salario_mensual")),
            "vacaciones_anuales": vac,
            "entra_roster": roster_norm(r.get("entra_roster")),
        })
    if not filas:
        raise HTTPException(400, "No se encontraron filas con dni.")

    with engine.begin() as conn:
        camp = conn.execute(text("select id from campanas where nombre = :n"), {"n": campana}).fetchone()
        if camp is None:
            # Deducir el tipo de país de la campaña según los agentes que se suben
            paises = set(f["pais"] for f in filas if f["pais"])
            tiene_es = "España" in paises
            tiene_co = "Colombia" in paises
            if tiene_es and tiene_co:
                ptipo = "MIXTA"
            elif tiene_es:
                ptipo = "ES"
            elif tiene_co:
                ptipo = "CO"
            else:
                ptipo = "MIXTA"
            camp = conn.execute(text(
                "insert into campanas (nombre, pais_tipo) values (:n, :pt) returning id"),
                {"n": campana, "pt": ptipo}).fetchone()
            campana_creada = True
        else:
            campana_creada = False
        id_camp = int(camp[0])

        # Si la campaña es nueva, clonar las tipificaciones marcadas como plantilla
        if campana_creada:
            ya = conn.execute(text(
                "select count(*) from crm_tipificaciones where campana_id = :c"), {"c": id_camp}).fetchone()
            if ya is None or int(ya[0]) == 0:
                conn.execute(text("""
                    insert into crm_tipificaciones (nombre, activo, orden, campana_id, es_plantilla)
                    select nombre, activo, orden, :c, false
                    from crm_tipificaciones
                    where es_plantilla = true and campana_id = 1
                """), {"c": id_camp})

            # Sembrar metas de adherencia por defecto para la campaña nueva
            metas_def = [
                ("adh_neta", "alto", 99, 99, 95, "%"),
                ("adh_bruta", "alto", 95, 95, 90, "%"),
                ("productividad", "alto", 96, 96, 90, "%"),
                ("utilizacion", "alto", 88, 88, 80, "%"),
                ("ocupacion", "alto", 85, 85, 75, "%"),
                ("tmo", "bajo", 300, 300, 360, "seg"),
            ]
            for ind, dirn, meta, verde, amar, uni in metas_def:
                conn.execute(text("""
                    insert into campana_metas (campana_id, indicador, direccion, meta, verde, amarillo, unidad)
                    values (:c, :i, :d, :m, :v, :a, :u)
                    on conflict (campana_id, indicador) do nothing
                """), {"c": id_camp, "i": ind, "d": dirn, "m": meta, "v": verde, "a": amar, "u": uni})

        existentes = {str(x[0]).strip().upper(): int(x[1]) for x in
                      conn.execute(text("select dni, id from agentes where dni is not null")).fetchall()}
        nuevos = [f for f in filas if f["dni"] not in existentes]
        upd    = [f for f in filas if f["dni"] in existentes]

        cols = ["dni", "nombre", "centro", "pais", "modo", "turno", "login_acd",
                "fecha_alta", "jornada_horas", "salario_mensual", "vacaciones_anuales",
                "entra_roster"]
        LOTE = 500
        for k in range(0, len(nuevos), LOTE):
            lote = nuevos[k:k + LOTE]
            vals = []; params = {}
            for j, f in enumerate(lote):
                vals.append(f"(:dni{j}, :nombre{j}, :centro{j}, :pais{j}, :modo{j}, :turno{j}, "
                            f":login_acd{j}, cast(:fecha_alta{j} as date), cast(:jornada_horas{j} as numeric), "
                            f"cast(:salario_mensual{j} as numeric), cast(:vacaciones_anuales{j} as numeric), "
                            f"coalesce(cast(:entra_roster{j} as boolean), true), "
                            f":camp{j}, 'ACTIVO')")
                for c in cols: params[f"{c}{j}"] = f[c]
                params[f"camp{j}"] = id_camp
            conn.execute(text(
                "insert into agentes (dni, nombre, centro, pais, modo, turno, login_acd, fecha_alta, "
                "jornada_horas, salario_mensual, vacaciones_anuales, entra_roster, campana_id, estado) values "
                + ",".join(vals)), params)

        for k in range(0, len(upd), LOTE):
            lote = upd[k:k + LOTE]
            vals = []; params = {}
            for j, f in enumerate(lote):
                vals.append(f"(:dni{j}, :nombre{j}, :centro{j}, :pais{j}, :modo{j}, :turno{j}, "
                            f":login_acd{j}, cast(:fecha_alta{j} as date), cast(:jornada_horas{j} as numeric), "
                            f"cast(:salario_mensual{j} as numeric), cast(:vacaciones_anuales{j} as numeric), "
                            f"cast(:entra_roster{j} as boolean), :camp{j})")
                for c in cols: params[f"{c}{j}"] = f[c]
                params[f"camp{j}"] = id_camp
            conn.execute(text(
                "update agentes a set "
                "nombre = coalesce(v.nombre, a.nombre), centro = coalesce(v.centro, a.centro), "
                "pais = coalesce(v.pais, a.pais), modo = coalesce(v.modo, a.modo), "
                "turno = coalesce(v.turno, a.turno), login_acd = coalesce(v.login_acd, a.login_acd), "
                "fecha_alta = coalesce(v.fecha_alta, a.fecha_alta), "
                "jornada_horas = coalesce(v.jornada_horas, a.jornada_horas), "
                "salario_mensual = coalesce(v.salario_mensual, a.salario_mensual), "
                "vacaciones_anuales = coalesce(v.vacaciones_anuales, a.vacaciones_anuales), "
                "entra_roster = coalesce(v.entra_roster, a.entra_roster), "
                "campana_id = v.camp "
                "from (values " + ",".join(vals) + ") as "
                "v(dni, nombre, centro, pais, modo, turno, login_acd, fecha_alta, jornada_horas, "
                "salario_mensual, vacaciones_anuales, entra_roster, camp) "
                "where upper(a.dni) = v.dni"), params)

        # usuarios: crear fila de login para agentes nuevos y vincular faltantes
        conn.execute(text("""
            insert into usuarios (email, rol, agente_id, activo)
            select lower(a.dni) || '@agentes.wfm', 'agente', a.id, true
            from agentes a
            where a.dni is not null
              and not exists (select 1 from usuarios u where u.email = lower(a.dni) || '@agentes.wfm')
        """))
        conn.execute(text("""
            update usuarios u set agente_id = a.id
            from agentes a
            where u.agente_id is null and u.email = lower(a.dni) || '@agentes.wfm'
        """))

    return {"ok": True, "campana": campana, "campana_creada": campana_creada,
            "agentes_creados": len(nuevos), "agentes_actualizados": len(upd)}


# ============================================================
#  Parámetros de planeación por campaña + mes
# ============================================================
@app.get("/campana-parametros")
def campana_parametros_get(
    x_api_key: str = Header(None),
    campana: str = None,
    mes: str = None,
):
    """Devuelve los parámetros guardados y el resumen del último cálculo de una campaña+mes.
    Si no hay nada guardado, devuelve existe=False (el frontend muestra los campos en blanco/0)."""
    check_key(x_api_key)
    engine = get_engine()
    cid = _id_campana(engine, campana)
    if cid is None or not mes:
        return {"ok": True, "existe": False}
    row = pd.read_sql(text("""
        select aht, sla, asa, occ, utl, esp_max, largo, nda_obj, paciencia, absentismo, estructura,
               res_volumen, res_espana, res_colombia, res_presentes, res_en_nomina, calculado_en
        from campana_parametros where campana_id=:c and mes=:m
    """), engine, params={"c": cid, "m": mes})
    if row.empty:
        return {"ok": True, "existe": False}
    r = row.iloc[0]
    def _n(v):
        return None if pd.isna(v) else (float(v) if isinstance(v, float) else v)
    return {
        "ok": True, "existe": True,
        "parametros": {
            "aht": _n(r["aht"]), "sla": _n(r["sla"]), "asa": _n(r["asa"]), "occ": _n(r["occ"]),
            "utl": _n(r["utl"]), "esp_max": _n(r["esp_max"]), "largo": _n(r["largo"]),
            "nda_obj": _n(r["nda_obj"]), "paciencia": _n(r["paciencia"]),
            "absentismo": _n(r["absentismo"]), "estructura": r["estructura"],
        },
        "resumen": {
            "volumen": _n(r["res_volumen"]), "espana": _n(r["res_espana"]),
            "colombia": _n(r["res_colombia"]), "presentes": _n(r["res_presentes"]),
            "en_nomina": _n(r["res_en_nomina"]),
            "calculado_en": (None if pd.isna(r["calculado_en"]) else str(r["calculado_en"])),
        },
    }


@app.post("/campana-parametros")
async def campana_parametros_set(
    x_api_key: str = Header(None),
    campana: str = Form(...),
    mes: str = Form(...),
    aht: int = Form(None),
    sla: float = Form(None),
    asa: int = Form(None),
    occ: float = Form(None),
    utl: float = Form(None),
    esp_max: int = Form(None),
    largo: int = Form(None),
    nda_obj: float = Form(None),
    paciencia: int = Form(None),
    absentismo: float = Form(None),
    estructura: str = Form("mixto"),
):
    """Guarda (sin calcular) los parámetros que el usuario coloca para una campaña+mes."""
    check_key(x_api_key)
    engine = get_engine()
    cid = _id_campana(engine, campana)
    if cid is None:
        raise HTTPException(400, f"La campaña '{campana}' no existe.")
    with engine.begin() as conn:
        conn.execute(text("""
            insert into campana_parametros
              (campana_id, mes, aht, sla, asa, occ, utl, esp_max, largo, nda_obj,
               paciencia, absentismo, estructura, actualizado_en)
            values
              (:cid, :mes, :aht, :sla, :asa, :occ, :utl, :esp_max, :largo, :nda_obj,
               :paciencia, :absentismo, :estructura, now())
            on conflict (campana_id, mes) do update set
              aht=excluded.aht, sla=excluded.sla, asa=excluded.asa, occ=excluded.occ,
              utl=excluded.utl, esp_max=excluded.esp_max, largo=excluded.largo,
              nda_obj=excluded.nda_obj, paciencia=excluded.paciencia,
              absentismo=excluded.absentismo, estructura=excluded.estructura,
              actualizado_en=now()
        """), {"cid": cid, "mes": mes, "aht": aht, "sla": sla, "asa": asa, "occ": occ,
               "utl": utl, "esp_max": esp_max, "largo": largo, "nda_obj": nda_obj,
               "paciencia": paciencia, "absentismo": absentismo, "estructura": estructura})
    return {"ok": True, "campana": campana, "mes": mes, "guardado": True}


# ============================================================
#  Call Capacity: pronóstico vs capacidad de atención por día e intervalo
#  Compara parrilla REAL (asignaciones) vs PLAN ÓPTIMO (optimizador)
# ============================================================
def _capacidad_intervalo(agentes, aht, paciencia, carga=None, nda_obj=0.96):
    """Máximas llamadas/hora que 'agentes' pueden atender cumpliendo el NDA objetivo,
    usando el modelo de abandono del motor (nivel_atencion). La 'carga' interna del motor
    va en Erlangs (llamadas*AHT/3600); buscamos el volumen de llamadas máximo soportable."""
    if agentes <= 0:
        return 0
    paso = max(1, agentes)
    V = 0; mejor = 0
    tope = int(agentes * 3600 / aht * 2)
    while V < tope:
        V += paso
        carga_erlangs = V * aht / 3600.0
        if motor.nivel_atencion(agentes, carga_erlangs, aht, paciencia) >= nda_obj:
            mejor = V
        else:
            break
    return mejor


def _agentes_por_hora_parrilla(engine, campana_id, mes):
    """Cuenta agentes con turno 'trabajo' por (fecha, hora) desde asignaciones."""
    df = pd.read_sql(text("""
        SELECT fecha, hora_inicio, hora_fin
        FROM asignaciones
        WHERE campana_id=:c AND tipo='trabajo'
          AND hora_inicio IS NOT NULL AND hora_fin IS NOT NULL
          AND to_char(fecha,'YYYY-MM')=:m
    """), engine, params={"c": campana_id, "m": mes})
    # cuenta por (fecha, hora) expandiendo cada turno a sus horas
    from collections import defaultdict
    cont = defaultdict(int)
    for _, r in df.iterrows():
        try:
            hi = int(str(r["hora_inicio"])[:2]); hf = int(str(r["hora_fin"])[:2])
        except Exception:
            continue
        horas = []
        h = hi
        # recorre del inicio al fin, manejando cruce de medianoche
        for _ in range(24):
            if h == hf:
                break
            horas.append(h)
            h = (h + 1) % 24
        for hh in horas:
            cont[(pd.Timestamp(r["fecha"]).date(), hh)] += 1
    return cont


@app.post("/call-capacity")
async def call_capacity(
    x_api_key: str = Header(None),
    mes: str = Form(...),
    campana: str = Form("Endesa"),
    aht: int = Form(420),
    paciencia: int = Form(90),
    largo: int = Form(9),
    sla: float = Form(0.80),
    asa: int = Form(20),
    occ: float = Form(0.65),
    utl: float = Form(0.88),
    esp_max: int = Form(12),
    nda_obj: float = Form(0.96),
    estructura: str = Form("mixto"),
    ajuste_pct: float = Form(0.0),
):
    """Pronóstico de llamadas vs capacidad de atención por día e intervalo.
    Capacidad calculada con dos plantillas: la parrilla REAL (asignaciones del mes) y el PLAN ÓPTIMO."""
    check_key(x_api_key)
    engine = get_engine()
    cid = _id_campana(engine, campana)
    if cid is None:
        raise HTTPException(400, f"La campaña '{campana}' no existe.")

    # 1) Pronóstico por fecha+intervalo (motor, con % de ajuste)
    file_bytes = _historico_a_csv_bytes(engine, cid)
    ajuste = (ajuste_pct or 0.0) / 100.0
    pron = motor.largo_desde_historico(file_bytes, mes, "Nacional España", 4, 6, ajuste, mixto=False)
    pron["fecha"] = pd.to_datetime(pron["fecha"]).dt.date

    # 2) Plan óptimo -> agentes por (dow, hora)
    S = motor.dimension_roster(pron.rename(columns={}), aht, sla, asa, occ, utl, esp_max, largo,
                               nda_obj, paciencia, estructura)
    turnos = motor.turnos_dict(S, largo)
    plan_por_dow_hora = {(d, h): motor.cubierto(S, turnos, d, h) for d in range(7) for h in range(24)}

    # 3) Parrilla real -> agentes por (fecha, hora)
    real_cont = _agentes_por_hora_parrilla(engine, cid, mes)

    # 4) Armar tabla por día e intervalo
    filas = []
    cap_cache_real = {}; cap_cache_plan = {}
    for _, r in pron.iterrows():
        f = r["fecha"]; h = int(r["intervalo"]); vol = float(r["volumen"])
        dow = pd.Timestamp(f).dayofweek
        ag_real = real_cont.get((f, h), 0)
        ag_plan = plan_por_dow_hora.get((dow, h), 0)
        if ag_real not in cap_cache_real:
            cap_cache_real[ag_real] = _capacidad_intervalo(ag_real, aht, paciencia, nda_obj=nda_obj)
        if ag_plan not in cap_cache_plan:
            cap_cache_plan[ag_plan] = _capacidad_intervalo(ag_plan, aht, paciencia, nda_obj=nda_obj)
        cap_real_max = cap_cache_real[ag_real]            # capacidad real (sin topar)
        cap_plan_max = cap_cache_plan[ag_plan]
        cap_real = min(cap_real_max, round(vol))          # capacidad efectiva (topada a pronosticadas)
        cap_plan = min(cap_plan_max, round(vol))
        filas.append({
            "fecha": str(f), "intervalo": h, "pronosticadas": round(vol),
            "agentes_real": ag_real,
            "capacidad_real_max": cap_real_max, "capacidad_real": cap_real,
            "holgura_real": cap_real_max - round(vol),    # + sobra capacidad, - falta
            "deficit_real": round(vol) - cap_real,
            "agentes_plan": ag_plan,
            "capacidad_plan_max": cap_plan_max, "capacidad_plan": cap_plan,
            "holgura_plan": cap_plan_max - round(vol),
            "deficit_plan": round(vol) - cap_plan,
        })

    df = pd.DataFrame(filas)
    # resumen por día
    por_dia = []
    for f, g in df.groupby("fecha"):
        por_dia.append({
            "fecha": f,
            "pronosticadas": int(g["pronosticadas"].sum()),
            "capacidad_real_max": int(g["capacidad_real_max"].sum()),
            "capacidad_real": int(g["capacidad_real"].sum()),
            "capacidad_plan_max": int(g["capacidad_plan_max"].sum()),
            "capacidad_plan": int(g["capacidad_plan"].sum()),
            "holgura_real": int(g["capacidad_real_max"].sum() - g["pronosticadas"].sum()),
            "holgura_plan": int(g["capacidad_plan_max"].sum() - g["pronosticadas"].sum()),
        })

    return {"ok": True, "campana": campana, "mes": mes,
            "por_intervalo": filas, "por_dia": por_dia,
            "totales": {
                "pronosticadas": int(df["pronosticadas"].sum()),
                "capacidad_real_max": int(df["capacidad_real_max"].sum()),
                "capacidad_real": int(df["capacidad_real"].sum()),
                "capacidad_plan_max": int(df["capacidad_plan_max"].sum()),
                "capacidad_plan": int(df["capacidad_plan"].sum()),
            }}


# ============================================================
#  Exportar Planeación a Excel
#  - Previsión y Capacity en matriz HORAS(filas) × DÍAS DEL MES(columnas)
#  - Capacity = min(Erlang NDA, pronosticadas), parrilla y plan en hojas aparte
#  - Req vs Prog: 7 gráficas (Lun-Dom) en una hoja
#  - Ocupación y NDA con gráfica
# ============================================================
@app.post("/planeacion/export")
async def planeacion_export(
    x_api_key: str = Header(None),
    mes: str = Form(...),
    campana: str = Form("Endesa"),
    aht: int = Form(420),
    sla: float = Form(0.80),
    asa: int = Form(20),
    occ: float = Form(0.65),
    utl: float = Form(0.88),
    esp_max: int = Form(12),
    largo: int = Form(9),
    nda_obj: float = Form(0.96),
    paciencia: int = Form(90),
    estructura: str = Form("mixto"),
    absentismo: float = Form(0.15),
    ajuste_pct: float = Form(0.0),
    incluir_capacity: bool = Form(True),
):
    check_key(x_api_key)
    engine = get_engine()
    cid = _id_campana(engine, campana)
    if cid is None:
        raise HTTPException(400, f"La campaña '{campana}' no existe.")

    from openpyxl.chart import BarChart, LineChart, Reference
    from openpyxl.utils import get_column_letter
    import math as _m

    file_bytes = _historico_a_csv_bytes(engine, cid)
    ajuste = (ajuste_pct or 0.0) / 100.0
    largo_df = motor.largo_desde_historico(file_bytes, mes, "Nacional España", 4, 6, ajuste, mixto=False)
    largo_df["fecha"] = pd.to_datetime(largo_df["fecha"]).dt.date
    S = motor.dimension_roster(largo_df, aht, sla, asa, occ, utl, esp_max, largo,
                               nda_obj, paciencia, estructura)
    turnos = motor.turnos_dict(S, largo)
    NOM = ["Lunes", "Martes", "Miércoles", "Jueves", "Viernes", "Sábado", "Domingo"]
    HORAS = [f"{h:02d}h" for h in range(24)]
    fechas = sorted(largo_df["fecha"].unique())

    presentes = S["te"] + S["tc"]
    en_nomina = _m.ceil(presentes / (1 - absentismo)) if absentismo < 1 else presentes

    # ---------- Resumen ----------
    df_resumen = pd.DataFrame([
        {"Concepto": "Campaña", "Valor": campana},
        {"Concepto": "Mes", "Valor": mes},
        {"Concepto": "Ajuste de volumen aplicado", "Valor": f"{ajuste_pct:+.0f}%"},
        {"Concepto": "Volumen total proyectado", "Valor": int(S["total"])},
        {"Concepto": "Estructura", "Valor": estructura},
        {"Concepto": "Agentes España (L-V)", "Valor": int(S["te"])},
        {"Concepto": "Agentes Colombia (24/7)", "Valor": int(S["tc"])},
        {"Concepto": "Total presentes", "Valor": int(presentes)},
        {"Concepto": "En nómina (con absentismo)", "Valor": int(en_nomina)},
        {"Concepto": "AHT (s)", "Valor": aht}, {"Concepto": "SLA", "Valor": sla},
        {"Concepto": "ASA (s)", "Valor": asa}, {"Concepto": "OCC", "Valor": occ},
        {"Concepto": "UTL", "Valor": utl}, {"Concepto": "NDA objetivo", "Valor": nda_obj},
        {"Concepto": "Paciencia (s)", "Valor": paciencia},
        {"Concepto": "Largo turno (h)", "Valor": largo},
        {"Concepto": "Absentismo", "Valor": absentismo},
    ])

    # ---------- Plan de turnos ----------
    plan = []
    for k in sorted(S["xe"], key=int):
        t = int(k)
        plan.append({"País": "España", "Inicio": f"{t:02d}:00",
                     "Fin": f"{(t+largo)%24:02d}:00", "Cantidad": S["xe"][k], "Libres": "Sáb, Dom"})
    for k in sorted(S["xc"], key=lambda x: (int(x.split("_")[0]), int(x.split("_")[1]))):
        t, p = map(int, k.split("_"))
        o = sorted(motor.LIBRES_VIZ[p])
        plan.append({"País": "Colombia", "Inicio": f"{t:02d}:00",
                     "Fin": f"{(t+largo)%24:02d}:00", "Cantidad": S["xc"][k],
                     "Libres": f"{NOM[o[0]][:3]}, {NOM[o[1]][:3]}"})
    df_plan = pd.DataFrame(plan) if plan else pd.DataFrame([{"País": "-", "Inicio": "-", "Fin": "-", "Cantidad": 0, "Libres": "-"}])

    # ---------- Previsión MATRIZ: filas=HORA, columnas=DÍA DEL MES ----------
    volp = largo_df.set_index(["fecha", "intervalo"])["volumen"].to_dict()
    def matriz_horas_x_dias(valor_fn):
        """Construye dict {hora: {fecha: valor}} -> DataFrame (filas hora, cols fechas)."""
        data = {}
        for h in range(24):
            data[h] = {f: valor_fn(f, h) for f in fechas}
        df = pd.DataFrame.from_dict(data, orient="index")  # filas hora, cols fechas
        df.index.name = "Hora"
        df = df.reset_index()
        df["Hora"] = [f"{h:02d}h" for h in df["Hora"]]
        return df

    df_prev = matriz_horas_x_dias(lambda f, h: round(float(volp.get((f, h), 0))))

    # ---------- Req vs Prog por día (para las 7 gráficas) ----------
    # tabla larga: una columna Hora + por cada día Req y Prog
    rp_cols = {"Hora": HORAS}
    for d in range(7):
        rp_cols[f"{NOM[d]} Req"] = [S["peak"][d][h] for h in range(24)]
        rp_cols[f"{NOM[d]} Prog"] = [motor.cubierto(S, turnos, d, h) for h in range(24)]
    df_rp = pd.DataFrame(rp_cols)

    # ---------- Ocupación y NDA (matriz día×hora + promedio) ----------
    df_occ = pd.DataFrame([{**{"Día": NOM[d]}, **{HORAS[h]: round(S["occ"][d][h]*100,1) for h in range(24)}} for d in range(7)])
    df_occ_prom = pd.DataFrame([{"Hora": HORAS[h],
                                 "Ocupación % (prom)": round(sum(S["occ"][d][h] for d in range(7))/7*100,1),
                                 "Objetivo %": round(occ*100,1)} for h in range(24)])
    df_nda = pd.DataFrame([{**{"Día": NOM[d]}, **{HORAS[h]: round(S["nda"][d][h]*100,1) for h in range(24)}} for d in range(7)])
    df_nda_prom = pd.DataFrame([{"Hora": HORAS[h],
                                 "NDA % (prom)": round(sum(S["nda"][d][h] for d in range(7))/7*100,1),
                                 "Objetivo %": round(nda_obj*100,1)} for h in range(24)])

    # ---------- Capacity (matriz HORA × DÍA), topado a pronosticadas ----------
    df_cap_real = df_cap_plan = None
    if incluir_capacity:
        plan_por_dow_hora = {(d, h): motor.cubierto(S, turnos, d, h) for d in range(7) for h in range(24)}
        real_cont = _agentes_por_hora_parrilla(engine, cid, mes)
        cap_real_cache = {}; cap_plan_cache = {}
        def cap_real_for(ag):
            if ag not in cap_real_cache: cap_real_cache[ag] = _capacidad_intervalo(ag, aht, paciencia, nda_obj=nda_obj)
            return cap_real_cache[ag]
        def cap_plan_for(ag):
            if ag not in cap_plan_cache: cap_plan_cache[ag] = _capacidad_intervalo(ag, aht, paciencia, nda_obj=nda_obj)
            return cap_plan_cache[ag]

        def cap_real_fn(f, h):
            vol = round(float(volp.get((f, h), 0)))
            ag = real_cont.get((f, h), 0)
            return min(cap_real_for(ag), vol)   # capacidad no supera pronosticadas
        def cap_plan_fn(f, h):
            vol = round(float(volp.get((f, h), 0)))
            dow = pd.Timestamp(f).dayofweek
            ag = plan_por_dow_hora.get((dow, h), 0)
            return min(cap_plan_for(ag), vol)
        df_cap_real = matriz_horas_x_dias(cap_real_fn)
        df_cap_plan = matriz_horas_x_dias(cap_plan_fn)

    # =========== Escribir Excel ===========
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as xw:
        df_resumen.to_excel(xw, sheet_name="Resumen", index=False)
        df_plan.to_excel(xw, sheet_name="Plan de turnos", index=False)
        df_prev.to_excel(xw, sheet_name="Previsión", index=False)
        df_rp.to_excel(xw, sheet_name="Req vs Prog", index=False)
        df_occ.to_excel(xw, sheet_name="Ocupación", index=False)
        df_occ_prom.to_excel(xw, sheet_name="Ocupación", index=False, startrow=len(df_occ)+3)
        df_nda.to_excel(xw, sheet_name="NDA", index=False)
        df_nda_prom.to_excel(xw, sheet_name="NDA", index=False, startrow=len(df_nda)+3)
        if df_cap_real is not None:
            df_cap_real.to_excel(xw, sheet_name="Capacity parrilla", index=False)
            df_cap_plan.to_excel(xw, sheet_name="Capacity plan", index=False)
            # fila de totales por día al final de cada matriz capacity
            for sheet, dfm in (("Capacity parrilla", df_cap_real), ("Capacity plan", df_cap_plan)):
                ws = xw.sheets[sheet]
                tot_row = len(dfm) + 2  # +1 header, +1 para quedar debajo
                ws.cell(row=tot_row, column=1, value="TOTAL día")
                ncols = dfm.shape[1]
                for col in range(2, ncols + 1):
                    L = get_column_letter(col)
                    ws.cell(row=tot_row, column=col, value=f"=SUM({L}2:{L}{len(dfm)+1})")

        # ---- 7 gráficas Req vs Prog (una por día) estilo herramienta web ----
        # Requerido = barras grises | Programado = línea escalonada verde con marcadores
        from openpyxl.chart.marker import Marker
        from openpyxl.drawing.line import LineProperties
        from openpyxl.chart.shapes import GraphicalProperties
        from openpyxl.drawing.fill import PatternFillProperties, ColorChoice
        ws_rp = xw.sheets["Req vs Prog"]
        anchor_row = 2
        for d in range(7):
            col_req = 2 + 2*d
            col_prog = 3 + 2*d
            # Barras (Requerido)
            bar = BarChart(); bar.type = "col"; bar.title = f"{NOM[d]} — Req vs Prog"
            bar.height = 7.5; bar.width = 15
            data_req = Reference(ws_rp, min_col=col_req, max_col=col_req, min_row=1, max_row=25)
            cats = Reference(ws_rp, min_col=1, min_row=2, max_row=25)
            bar.add_data(data_req, titles_from_data=True); bar.set_categories(cats)
            # color gris para las barras
            try:
                bar.series[0].graphicalProperties.solidFill = "BFBFBF"
            except Exception:
                pass
            # Línea (Programado) escalonada verde con marcadores
            line = LineChart()
            data_prog = Reference(ws_rp, min_col=col_prog, max_col=col_prog, min_row=1, max_row=25)
            line.add_data(data_prog, titles_from_data=True)
            s = line.series[0]
            s.smooth = False          # sin suavizado
            try:
                s.graphicalProperties = GraphicalProperties()
                s.graphicalProperties.line = LineProperties(solidFill="2E7D32", w=20000)  # verde
                s.marker = Marker(symbol="circle", size=5)
                s.marker.graphicalProperties = GraphicalProperties(solidFill="2E7D32")
            except Exception:
                pass
            bar.y_axis.title = None; bar.x_axis.delete = False
            bar += line  # combinar barras + línea en el mismo gráfico
            anchor_col_letter = get_column_letter(df_rp.shape[1] + 2 + (d % 2) * 9)
            anchor = f"{anchor_col_letter}{anchor_row + (d // 2) * 16}"
            ws_rp.add_chart(bar, anchor)

        # ---- Gráfica volumen total por día (Previsión) ----
        ws_pv = xw.sheets["Previsión"]
        ncol = df_prev.shape[1]  # Hora + N días
        tot_row = 26  # debajo de las 24 horas (fila 25) dejamos fila total
        ws_pv.cell(row=tot_row, column=1, value="TOTAL día")
        for col in range(2, ncol + 1):
            L = get_column_letter(col)
            ws_pv.cell(row=tot_row, column=col, value=f"=SUM({L}2:{L}25)")
        chb = BarChart(); chb.title = "Volumen total por día"; chb.height = 9; chb.width = 26
        dref = Reference(ws_pv, min_col=2, max_col=ncol, min_row=tot_row, max_row=tot_row)
        chb.add_data(dref, titles_from_data=False)
        ws_pv.add_chart(chb, f"A{tot_row + 2}")

        # ---- Gráfica Ocupación ----
        ws_oc = xw.sheets["Ocupación"]
        r0 = len(df_occ) + 4
        cho = BarChart(); cho.title = "Ocupación por hora (% prom) vs objetivo"; cho.height = 9; cho.width = 20
        do = Reference(ws_oc, min_col=2, max_col=3, min_row=r0, max_row=r0+24)
        co = Reference(ws_oc, min_col=1, min_row=r0+1, max_row=r0+24)
        cho.add_data(do, titles_from_data=True); cho.set_categories(co)
        ws_oc.add_chart(cho, f"A{r0+27}")

        # ---- Gráfica NDA ----
        ws_nd = xw.sheets["NDA"]
        rn = len(df_nda) + 4
        chn = BarChart(); chn.title = "NDA por hora (% prom) vs objetivo"; chn.height = 9; chn.width = 20
        dn = Reference(ws_nd, min_col=2, max_col=3, min_row=rn, max_row=rn+24)
        cn = Reference(ws_nd, min_col=1, min_row=rn+1, max_row=rn+24)
        chn.add_data(dn, titles_from_data=True); chn.set_categories(cn)
        ws_nd.add_chart(chn, f"A{rn+27}")
    buf.seek(0)

    from fastapi.responses import StreamingResponse
    nombre = f"planeacion_{campana}_{mes}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{nombre}"'},
    )


# ============================================================
#  Tipificaciones · clonar plantilla a una campaña existente
# ============================================================
@app.post("/tipificaciones/clonar-plantilla")
async def tipificaciones_clonar_plantilla(
    x_api_key: str = Header(None),
    campana: str = Form(...),
    solo_si_vacia: bool = Form(True),
):
    """Copia las tipificaciones marcadas como plantilla (es_plantilla=true) a la campaña indicada.
    Por defecto solo lo hace si la campaña aún no tiene tipificaciones propias."""
    check_key(x_api_key)
    engine = get_engine()
    cid = _id_campana(engine, campana)
    if cid is None:
        raise HTTPException(400, f"La campaña '{campana}' no existe.")
    with engine.begin() as conn:
        if solo_si_vacia:
            ya = conn.execute(text(
                "select count(*) from crm_tipificaciones where campana_id = :c"), {"c": cid}).fetchone()
            if ya and int(ya[0]) > 0:
                return {"ok": True, "campana": campana, "clonadas": 0, "motivo": "ya tenía tipificaciones"}
        res = conn.execute(text("""
            insert into crm_tipificaciones (nombre, activo, orden, campana_id, es_plantilla)
            select nombre, activo, orden, :c, false
            from crm_tipificaciones
            where es_plantilla = true and campana_id = 1
        """), {"c": cid})
        n = res.rowcount or 0
    return {"ok": True, "campana": campana, "clonadas": n}
